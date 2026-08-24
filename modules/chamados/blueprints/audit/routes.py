from __future__ import annotations
import json
from flask import render_template, request, jsonify, send_file
from flask_login import login_required
from sqlalchemy import desc
from modules.chamados.models import AuditLog
from . import audit_bp


from dataclasses import dataclass
from typing import Iterable

@dataclass(frozen=True)
class ActionGroup:
    key: str
    label: str
    actions: tuple[str, ...]

ACTION_GROUPS: tuple[ActionGroup, ...] = (
    ActionGroup("create", "Inclusões", ("create", "link", "upload")),
    ActionGroup("update", "Edições", ("update", "move", "assign", "status", "reply")),
    ActionGroup("delete", "Exclusões", ("delete", "unlink")),
)


def _parse_int(v, default):
    try:
        v = int(v)
        return v if v > 0 else default
    except Exception:
        return default


def _actions_for_group(key: str) -> Iterable[str]:
    for group in ACTION_GROUPS:
        if group.key == key:
            return group.actions
    return ()


def _decode_payload(raw: str | None):
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None

@audit_bp.route("/", methods=["GET"])
@login_required
def page():
    q_entity = (request.args.get("entity_type") or "").strip()
    q_action = (request.args.get("action") or "").strip()
    q_actor  = (request.args.get("actor") or "").strip()
    q_text   = (request.args.get("q") or "").strip()

    page_num = _parse_int(request.args.get("page", 1), 1)
    per_page = min(100, _parse_int(request.args.get("per_page", 25), 25))

    qry = AuditLog.query
    if q_entity:
        qry = qry.filter(AuditLog.entity_type == q_entity)
    if q_action:
        qry = qry.filter(AuditLog.action == q_action)
    if q_actor:
        like = f"%{q_actor}%"
        qry = qry.filter(
            (AuditLog.actor_email.ilike(like)) |
            (AuditLog.actor_name.ilike(like)) |
            (AuditLog.actor_id == q_actor)
        )
    if q_text:
        like = f"%{q_text}%"
        qry = qry.filter(AuditLog.message.ilike(like))

    qry = qry.order_by(desc(AuditLog.created_at))
    pagination = qry.paginate(page=page_num, per_page=per_page, error_out=False)
    rows = pagination.items
    for row in rows:
        row.before_data = _decode_payload(row.before)
        row.after_data = _decode_payload(row.after)

    return render_template(
        "chamados/audit/index.html",
        rows=rows,
        pagination=pagination,
        filters=dict(entity_type=q_entity, action=q_action, actor=q_actor, q=q_text, per_page=per_page),
    )

@audit_bp.route("/api", methods=["GET"])
@login_required
def api_list():
    q_entity = (request.args.get("entity_type") or "").strip()
    q_action = (request.args.get("action") or "").strip()
    limit = _parse_int(request.args.get("limit", 50), 50)
    limit = min(200, max(1, limit))

    qry = AuditLog.query
    if q_entity:
        qry = qry.filter(AuditLog.entity_type == q_entity)
    if q_action:
        qry = qry.filter(AuditLog.action == q_action)
    qry = qry.order_by(desc(AuditLog.created_at)).limit(limit)

    def dump(r: AuditLog):
        return dict(
            id=r.id,
            created_at=r.created_at.isoformat() if r.created_at else None,
            actor=dict(id=r.actor_id, email=r.actor_email, name=r.actor_name),
            ip=r.ip, ua=r.ua,
            entity_type=r.entity_type, entity_id=r.entity_id,
            action=r.action,
            message=r.message,
            before=_decode_payload(r.before) or r.before,
            after=_decode_payload(r.after) or r.after,
        )

    return jsonify([dump(r) for r in qry.all()])


@audit_bp.route("/export", methods=["GET"])
@login_required
def export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime

    q_entity = (request.args.get("entity_type") or "").strip()
    q_action = (request.args.get("action") or "").strip()
    q_actor = (request.args.get("actor") or "").strip()
    q_text = (request.args.get("q") or "").strip()
    q_operation = (request.args.get("operation") or "").strip()

    qry = AuditLog.query
    if q_entity:
        qry = qry.filter(AuditLog.entity_type == q_entity)
    if q_action:
        qry = qry.filter(AuditLog.action == q_action)
    elif q_operation:
        actions = list(_actions_for_group(q_operation))
        if actions:
            qry = qry.filter(AuditLog.action.in_(actions))
    if q_actor:
        like = f"%{q_actor}%"
        qry = qry.filter(
            (AuditLog.actor_email.ilike(like))
            | (AuditLog.actor_name.ilike(like))
            | (AuditLog.actor_id == q_actor)
        )
    if q_text:
        like = f"%{q_text}%"
        qry = qry.filter(AuditLog.message.ilike(like))

    qry = qry.order_by(desc(AuditLog.created_at))
    logs = qry.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B3B8C", end_color="0B3B8C", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "ID", "Data/Hora", "Autor (Email)", "Autor (Nome)", 
        "IP", "Entity Type", "Entity ID", "Action", "Mensagem"
    ]
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    for log in logs:
        created_str = log.created_at.strftime("%d/%m/%Y %H:%M:%S") if log.created_at else ""
        row_data = [
            log.id,
            created_str,
            log.actor_email,
            log.actor_name,
            log.ip,
            log.entity_type,
            log.entity_id,
            log.action,
            log.message
        ]
        ws.append(row_data)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            if col_idx in (1, 2, 5, 7, 8):
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


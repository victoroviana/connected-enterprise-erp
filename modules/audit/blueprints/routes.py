from __future__ import annotations

import json
from dataclasses import dataclass
from typing import Iterable

from flask import jsonify, render_template, request, send_file
from flask_login import login_required
from sqlalchemy import desc

from modules.audit.models import AuditLog

from . import audit_bp


@dataclass(frozen=True)
class ActionGroup:
    key: str
    label: str
    actions: tuple[str, ...]


ACTION_GROUPS: tuple[ActionGroup, ...] = (
    ActionGroup("create", "Incluses", ("create", "link", "upload")),
    ActionGroup("update", "Edies", ("update", "move", "assign", "status", "reply")),
    ActionGroup("delete", "Excluses", ("delete", "unlink")),
)


def _parse_int(raw: str | None, default: int) -> int:
    try:
        value = int(raw)
        return value if value > 0 else default
    except (TypeError, ValueError):
        return default


def _actions_for_group(key: str) -> Iterable[str]:
    for group in ACTION_GROUPS:
        if group.key == key:
            return group.actions
    return ()


def _decode_payload(raw: str | None):
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


@audit_bp.route("/", methods=["GET"])
@login_required
def page():
    q_entity = (request.args.get("entity_type") or "").strip()
    q_action = (request.args.get("action") or "").strip()
    q_actor = (request.args.get("actor") or "").strip()
    q_text = (request.args.get("q") or "").strip()
    q_operation = (request.args.get("operation") or "").strip()

    page_num = _parse_int(request.args.get("page"), 1)
    per_page = min(100, _parse_int(request.args.get("per_page"), 25))

    qry = AuditLog.query
    if q_entity:
        qry = qry.filter(AuditLog.entity_type == q_entity)
    if q_action:
        qry = qry.filter(AuditLog.action == q_action)
    elif q_operation:
        actions = list(_actions_for_group(q_operation))
        if actions:
            qry = qry.filter(AuditLog.action.in_(actions))
    if q_actor:
        like = f"%{q_actor}%"
        qry = qry.filter(
            (AuditLog.actor_email.ilike(like))
            | (AuditLog.actor_name.ilike(like))
            | (AuditLog.actor_id == q_actor)
        )
    if q_text:
        like = f"%{q_text}%"
        qry = qry.filter(AuditLog.message.ilike(like))

    qry = qry.order_by(desc(AuditLog.created_at))
    pagination = qry.paginate(page=page_num, per_page=per_page, error_out=False)
    rows = pagination.items
    for row in rows:
        row.before_data = _decode_payload(row.before)
        row.after_data = _decode_payload(row.after)

    filters = dict(
        entity_type=q_entity,
        action=q_action,
        actor=q_actor,
        q=q_text,
        operation=q_operation,
        per_page=per_page,
    )

    return render_template(
        "audit/index.html",
        rows=rows,
        pagination=pagination,
        filters=filters,
        action_groups=ACTION_GROUPS,
    )


@audit_bp.route("/api", methods=["GET"])
@login_required
def api_list():
    q_entity = (request.args.get("entity_type") or "").strip()
    q_entity_id = (request.args.get("entity_id") or "").strip()
    q_action = (request.args.get("action") or "").strip()
    q_operation = (request.args.get("operation") or "").strip()

    limit = _parse_int(request.args.get("limit"), 50)
    limit = min(200, max(1, limit))

    qry = AuditLog.query
    if q_entity:
        qry = qry.filter(AuditLog.entity_type == q_entity)
    if q_entity_id:
        qry = qry.filter(AuditLog.entity_id == q_entity_id)
    if q_action:
        qry = qry.filter(AuditLog.action == q_action)
    elif q_operation:
        actions = list(_actions_for_group(q_operation))
        if actions:
            qry = qry.filter(AuditLog.action.in_(actions))

    qry = qry.order_by(desc(AuditLog.created_at)).limit(limit)

    def dump(row: AuditLog) -> dict[str, object | None]:
        return dict(
            id=row.id,
            created_at=row.created_at.isoformat() if row.created_at else None,
            actor=dict(id=row.actor_id, email=row.actor_email, name=row.actor_name),
            ip=row.ip,
            ua=row.ua,
            entity_type=row.entity_type,
            entity_id=row.entity_id,
            action=row.action,
            message=row.message,
            before=_decode_payload(row.before) or row.before,
            after=_decode_payload(row.after) or row.after,
        )

    return jsonify([dump(r) for r in qry.all()])


@audit_bp.route("/export", methods=["GET"])
@login_required
def export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime

    q_entity = (request.args.get("entity_type") or "").strip()
    q_action = (request.args.get("action") or "").strip()
    q_actor = (request.args.get("actor") or "").strip()
    q_text = (request.args.get("q") or "").strip()
    q_operation = (request.args.get("operation") or "").strip()

    qry = AuditLog.query
    if q_entity:
        qry = qry.filter(AuditLog.entity_type == q_entity)
    if q_action:
        qry = qry.filter(AuditLog.action == q_action)
    elif q_operation:
        actions = list(_actions_for_group(q_operation))
        if actions:
            qry = qry.filter(AuditLog.action.in_(actions))
    if q_actor:
        like = f"%{q_actor}%"
        qry = qry.filter(
            (AuditLog.actor_email.ilike(like))
            | (AuditLog.actor_name.ilike(like))
            | (AuditLog.actor_id == q_actor)
        )
    if q_text:
        like = f"%{q_text}%"
        qry = qry.filter(AuditLog.message.ilike(like))

    qry = qry.order_by(desc(AuditLog.created_at))
    logs = qry.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B3B8C", end_color="0B3B8C", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "ID", "Data/Hora", "Autor (Email)", "Autor (Nome)", 
        "IP", "Entity Type", "Entity ID", "Action", "Mensagem"
    ]
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    for log in logs:
        created_str = log.created_at.strftime("%d/%m/%Y %H:%M:%S") if log.created_at else ""
        row_data = [
            log.id,
            created_str,
            log.actor_email,
            log.actor_name,
            log.ip,
            log.entity_type,
            log.entity_id,
            log.action,
            log.message
        ]
        ws.append(row_data)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            if col_idx in (1, 2, 5, 7, 8):
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

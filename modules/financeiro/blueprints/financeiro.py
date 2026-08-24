"""Financeiro blueprint: contas a receber e cota mensal/trimestral."""
from __future__ import annotations

from datetime import date, datetime
import calendar
import html
import io
import math
import re
import smtplib
import threading
import unicodedata
from email.message import EmailMessage
from email.utils import formataddr, parseaddr
from typing import Any

from utils.helpers import (
    normalize_dept_name as _normalize_dept_name,
    wants_json as _wants_json,
    paginate as _paginate,
    format_date as _format_date,
    format_datetime as _format_datetime,
    submit_bg_task,
)

import requests
from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import current_user
from sqlalchemy import bindparam, text

from extensions import db
from modules.audit.utils import write_audit_external
from modules.propostas.blueprints.auth import login_required
from modules.propostas.blueprints.auth.permissions_utils import normalize_permissions
from modules.propostas.services.proposal_email import _resolve_mail_settings
from modules.propostas.gerar_proposta import render_proposta_html_pdf


financeiro_bp = Blueprint("financeiro_bp", __name__, url_prefix="/financeiro")

RESULTS_PER_PAGE = 10

FINANCEIRO_UNIDADES_COTA = [
    "SOLLUS RJ",
    "TECNHOSOLLUS RJ",
    "TECNHOSOLLUS ES",
    "SS SANTOS",
    "SOLLUS PR",
    "SOLLUS SP",
]
FINANCEIRO_ALLOWED_DEPTS = {"FINANCEIRO", "FINANCAS"}

FINANCEIRO_CONTAS_ENDPOINTS = {
    "financeiro_bp.contas_receber",
    "financeiro_bp.contas_receber_adicionar",
    "financeiro_bp.contas_receber_subpendencia",
    "financeiro_bp.contas_receber_subpendencias",
    "financeiro_bp.contas_receber_editar",
    "financeiro_bp.contas_receber_editar_sub",
    "financeiro_bp.contas_receber_quitar",
    "financeiro_bp.contas_receber_update_info",
    "financeiro_bp.contas_receber_historico",
    "financeiro_bp.contas_receber_valor",
    "financeiro_bp.contas_receber_bloqueio",
    "financeiro_bp.contas_receber_cancelamento",
    "financeiro_bp.contas_receber_deferimento",
    "financeiro_bp.contas_receber_excluir",
    "financeiro_bp.verifica_empresa",
}
FINANCEIRO_CANCELADOS_ENDPOINTS = {
    "financeiro_bp.contas_receber_cancelados",
}
FINANCEIRO_COTA_ENDPOINTS = {
    "financeiro_bp.cota_dashboard",
    "financeiro_bp.cota_adicionar",
    "financeiro_bp.cota_valor",
    "financeiro_bp.cota_valor_pago",
    "financeiro_bp.cota_trimestre",
    "financeiro_bp.cota_pdf_mensal",
    "financeiro_bp.cota_pdf_trimestre",
    "financeiro_bp.cota_enviar_email",
    "financeiro_bp.cota_fechar_mes",
}

MESES_PT = {
    "January": "Janeiro",
    "February": "Fevereiro",
    "March": "Março",
    "April": "Abril",
    "May": "Maio",
    "June": "Junho",
    "July": "Julho",
    "August": "Agosto",
    "September": "Setembro",
    "October": "Outubro",
    "November": "Novembro",
    "December": "Dezembro",
}




def _dept_is_financeiro() -> bool:
    try:
        names = list(getattr(current_user, "department_names", []) or [])
    except Exception:
        names = []
    if not names:
        try:
            dept = getattr(current_user, "department", None)
            name = getattr(dept, "name", "") if dept else ""
            if name:
                names = [name]
        except Exception:
            names = []
    return any(_normalize_dept_name(name) in FINANCEIRO_ALLOWED_DEPTS for name in names)


def _is_admin_user() -> bool:
    role = (getattr(current_user, "tipo", None) or session.get("tipo") or "").lower()
    return role == "admin"


def _has_finance_permission() -> bool:
    perms = getattr(current_user, "permissions", None) or {}
    try:
        perms = normalize_permissions(perms if isinstance(perms, dict) else {})
    except Exception:
        perms = perms if isinstance(perms, dict) else {}
    return any(
        perms.get(key, False)
        for key in ("financeiro", "financeiro_contas", "financeiro_cancelados", "financeiro_cota")
    )


@financeiro_bp.before_request
def _check_financeiro_permissions():
    from flask import request
    if "/api/" in getattr(request, "path", ""):
        return
    endpoint = getattr(request, "endpoint", "") or ""
    if endpoint and not endpoint.startswith("financeiro_bp."):
        return
    if endpoint == "financeiro_bp.sem_permissao":
        return
    if not current_user.is_authenticated:
        return
    if _is_admin_user() or _dept_is_financeiro():
        return
    if _has_finance_permission() and endpoint in FINANCEIRO_CONTAS_ENDPOINTS:
        return
    if _has_finance_permission() and endpoint in FINANCEIRO_CANCELADOS_ENDPOINTS:
        return
    if _has_finance_permission() and endpoint in FINANCEIRO_COTA_ENDPOINTS:
        return
    if _wants_json():
        return jsonify({"ok": False, "message": "Você não tem permissão para acessar o financeiro."}), 403
    flash(
        "Você não tem permissão para acessar o Financeiro. Procure seu superior caso precise de acesso.",
        "warning",
    )
    return redirect(url_for("financeiro_bp.sem_permissao"))


@financeiro_bp.route("/sem-permissao")
@login_required
def sem_permissao():
    return render_template("errors/403.html", area_label="o Financeiro")





def _safe_int(value: Any, default: int = 1) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_money(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return 0.0
    raw = raw.replace("R$", "").replace(" ", "")
    raw = raw.replace(".", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _format_money(value: Any) -> str:
    try:
        parsed = float(value or 0.0)
    except (TypeError, ValueError):
        parsed = 0.0
    return f"{parsed:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")






def _parse_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _now_br_string() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def _calc_dias_atraso(value: Any) -> int:
    dt = _parse_date(value)
    if not dt:
        return 0
    today = date.today()
    return max(0, (today - dt).days)


def _current_user_name() -> str:
    try:
        if current_user and getattr(current_user, "is_authenticated", False):
            return (
                getattr(current_user, "nome_completo", None)
                or getattr(current_user, "usuario", None)
                or getattr(current_user, "email", None)
                or "Sistema"
            )
    except Exception:
        pass
    return (
        session.get("logged_in_user")
        or session.get("usuario_nome")
        or session.get("nome")
        or "Sistema"
    )


def _row_to_dict(row: Any) -> dict[str, Any]:
    mapping = getattr(row, "_mapping", None)
    data = dict(mapping) if mapping is not None else dict(row)
    return {key: _sanitize_mojibake(value) for key, value in data.items()}


def _sanitize_mojibake(value: Any) -> Any:
    if value is None:
        return value
    if isinstance(value, str):
        try:
            fixed = value.encode("latin1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return value
        return fixed
    if isinstance(value, (list, tuple)):
        fixed = [_sanitize_mojibake(item) for item in value]
        return type(value)(fixed)
    return value





def _fetch_contas(
    *,
    statuses: list[str],
    search: str,
    date_start: str,
    date_end: str,
    page: int,
    per_page: int,
    order: str,
) -> tuple[list[dict[str, Any]], int]:
    columns = (
        "id, cliente, cnpj, contrato, software, data_primeira_pendencia, qt_pendencias, "
        "dias_atraso, data_bloqueio, cancelamento, deferimento_cancelamento, informacoes, "
        "empresa_responsavel, valor, id_pai, total, status"
    )
    where = ["id_pai = 0"]
    params: dict[str, Any] = {}

    if statuses:
        where.append("status IN :statuses")
        params["statuses"] = statuses
    if search:
        where.append("(cnpj LIKE :search OR cliente LIKE :search)")
        params["search"] = f"%{search}%"
    if date_start:
        where.append("data_primeira_pendencia >= :date_start")
        params["date_start"] = date_start
    if date_end:
        where.append("data_primeira_pendencia <= :date_end")
        params["date_end"] = date_end

    where_sql = " AND ".join(where)
    count_sql = text(f"SELECT COUNT(id) AS total FROM contas_receber WHERE {where_sql}").bindparams(
        bindparam("statuses", expanding=True)
    )
    total = db.session.execute(count_sql, params).scalar() or 0

    limit = per_page if per_page else 10
    offset = (page - 1) * limit
    query = text(
        f"SELECT {columns} FROM contas_receber WHERE {where_sql} "
        f"ORDER BY data_primeira_pendencia {order} LIMIT :limit OFFSET :offset"
    ).bindparams(bindparam("statuses", expanding=True))

    params.update({"limit": limit, "offset": offset})
    rows = db.session.execute(query, params).fetchall()
    payload: list[dict[str, Any]] = []
    for row in rows:
        item = _row_to_dict(row)
        item["data_primeira_pendencia_br"] = _format_date(item.get("data_primeira_pendencia"))
        item["data_bloqueio_br"] = _format_datetime(item.get("data_bloqueio"))
        item["cancelamento_br"] = _format_datetime(item.get("cancelamento"))
        item["deferimento_br"] = _format_datetime(item.get("deferimento_cancelamento"))
        item["valor_display"] = _format_money(item.get("valor"))
        item["total_display"] = _format_money(item.get("total"))
        item["dias_atraso_calc"] = _calc_dias_atraso(item.get("data_primeira_pendencia"))
        payload.append(item)
    return payload, total


def _build_url(endpoint: str, **updates: Any) -> str:
    args = dict(request.args)
    for key, value in updates.items():
        if value in (None, ""):
            args.pop(key, None)
        else:
            args[key] = value
    return url_for(endpoint, **args)


def _resolve_finance_smtp_settings() -> dict[str, Any]:
    base = _resolve_mail_settings()
    name, email = parseaddr(base.get("sender", "") or "")
    return {
        "host": base["host"],
        "port": base["port"],
        "username": base.get("username"),
        "password": base.get("password"),
        "use_ssl": base.get("use_ssl", False),
        "use_tls": base.get("use_tls", False),
        "from_email": email or base["sender"],
        "from_name": name or "Financeiro Sollus",
        "reply_to": base.get("reply_to"),
    }


def _resolve_cota_smtp_settings() -> dict[str, Any]:
    cfg = current_app.config
    defaults = {
        "host": "smtp.sollustecnologia.com",
        "port": 587,
        "username": "contratos.automatico@sollustecnologia.com",
        "password": None,  # Must be configured via COTA_SMTP_PASSWORD or MAIL_PASSWORD env var
        "from_email": "automatico@sollustecnologia.com",
        "from_name": "Sollus Tecnologia - Cota Geral",
        "to": ["automatico.cota@sollustecnologia.com"],
        "cc": [],
    }
    return {
        "host": cfg.get("COTA_SMTP_HOST") or cfg.get("MAIL_SERVER") or defaults["host"],
        "port": int(cfg.get("COTA_SMTP_PORT") or cfg.get("MAIL_PORT") or defaults["port"]),
        "username": cfg.get("COTA_SMTP_USERNAME") or cfg.get("MAIL_USERNAME") or defaults["username"],
        "password": cfg.get("COTA_SMTP_PASSWORD") or cfg.get("MAIL_PASSWORD") or defaults["password"],
        "from_email": cfg.get("COTA_FROM_EMAIL") or cfg.get("MAIL_DEFAULT_SENDER") or defaults["from_email"],
        "from_name": cfg.get("COTA_FROM_NAME") or defaults["from_name"],
        "to": cfg.get("COTA_EMAIL_TO") or defaults["to"],
        "cc": cfg.get("COTA_EMAIL_CC") or defaults["cc"],
        "use_tls": cfg.get("COTA_SMTP_USE_TLS") if cfg.get("COTA_SMTP_USE_TLS") is not None else cfg.get("MAIL_USE_TLS", True),
        "use_ssl": cfg.get("COTA_SMTP_USE_SSL") if cfg.get("COTA_SMTP_USE_SSL") is not None else cfg.get("MAIL_USE_SSL", False),
    }


def _send_email(settings: dict[str, Any], subject: str, html_body: str) -> bool:
    if current_app.config.get("MAIL_ENABLED", True) is False:
        try:
            write_audit_external(
                entity_type="financeiro_email",
                action="email_skip",
                message="Envio de email do financeiro ignorado: MAIL_ENABLED falso.",
                after={"assunto": subject, "status": "disabled"},
            )
        except Exception:
            current_app.logger.exception("Falha ao auditar envio de email do financeiro (skip).")
        return False

    recipients = settings.get("to") or []
    if isinstance(recipients, str):
        recipients = [recipients]
    cc_list = settings.get("cc") or []
    if isinstance(cc_list, str):
        cc_list = [cc_list]

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = formataddr((settings.get("from_name"), settings.get("from_email")))
    msg["To"] = ", ".join(recipients)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg.set_content(html_body.replace("<br>", "\n").replace("<br/>", "\n"))
    msg.add_alternative(html_body, subtype="html")

    import ssl
    skip_tls_verify = str(current_app.config.get("MAIL_SKIP_TLS_VERIFY", "")).strip() == "1"
    context = ssl.create_default_context()
    if skip_tls_verify:
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE

    try:
        use_ssl = settings.get("use_ssl", False) or settings["port"] == 465
        if use_ssl:
            smtp_client = smtplib.SMTP_SSL(settings["host"], settings["port"], context=context, timeout=30)
        else:
            smtp_client = smtplib.SMTP(settings["host"], settings["port"], timeout=30)

        with smtp_client as smtp:
            smtp.ehlo()
            if not use_ssl and settings.get("use_tls"):
                smtp.starttls(context=context)
                smtp.ehlo()
            if settings.get("username"):
                smtp.login(settings["username"], settings.get("password") or "")
            smtp.send_message(msg)
        try:
            write_audit_external(
                entity_type="financeiro_email",
                action="email_send",
                message="Envio de email do financeiro concluido.",
                after={"assunto": subject, "to": recipients, "cc": cc_list, "status": "success"},
            )
        except Exception:
            current_app.logger.exception("Falha ao auditar envio de email do financeiro.")
        return True
    except Exception:
        current_app.logger.exception("Falha ao enviar e-mail do financeiro")
        try:
            write_audit_external(
                entity_type="financeiro_email",
                action="email_error",
                message="Falha ao enviar email do financeiro.",
                after={"assunto": subject, "to": recipients, "cc": cc_list, "status": "error"},
            )
        except Exception:
            current_app.logger.exception("Falha ao auditar erro no email do financeiro.")
        return False


def _dispatch_email_async(settings: dict[str, Any], subject: str, html_body: str) -> None:
    app = current_app._get_current_object()
    submit_bg_task(app, _send_email, settings, subject, html_body)


@financeiro_bp.route("/contas-receber")
@login_required
def contas_receber():
    page = _safe_int(request.args.get("page"), 1)
    search = (request.args.get("search") or "").strip()
    date_start = (request.args.get("date_start") or "").strip()
    date_end = (request.args.get("date_end") or "").strip()

    contas, total = _fetch_contas(
        statuses=["ABERTO"],
        search=search,
        date_start=date_start,
        date_end=date_end,
        page=page,
        per_page=RESULTS_PER_PAGE,
        order="ASC",
    )
    pagination = _paginate(total, page, RESULTS_PER_PAGE)
    stats = [
        {"label": "Pendências", "value": total},
        {"label": "Página", "value": f"{pagination['page']}/{pagination['pages']}"},
    ]

    return render_template(
        "admin/financeiro/contas_receber.html",
        page_mode="aberto",
        contas=contas,
        pagination=pagination,
        stats=stats,
        search_value=search,
        date_start=date_start,
        date_end=date_end,
        build_url=lambda **kw: _build_url("financeiro_bp.contas_receber", **kw),
    )


@financeiro_bp.route("/contas-receber/cancelados")
@login_required
def contas_receber_cancelados():
    page = _safe_int(request.args.get("page"), 1)
    search = (request.args.get("search") or "").strip()
    date_start = (request.args.get("date_start") or "").strip()
    date_end = (request.args.get("date_end") or "").strip()

    contas, total = _fetch_contas(
        statuses=["FECHADO", "QUITADO"],
        search=search,
        date_start=date_start,
        date_end=date_end,
        page=page,
        per_page=RESULTS_PER_PAGE,
        order="DESC",
    )
    pagination = _paginate(total, page, RESULTS_PER_PAGE)
    stats = [
        {"label": "Registros", "value": total},
        {"label": "Página", "value": f"{pagination['page']}/{pagination['pages']}"},
    ]

    return render_template(
        "admin/financeiro/contas_receber.html",
        page_mode="cancelados",
        contas=contas,
        pagination=pagination,
        stats=stats,
        search_value=search,
        date_start=date_start,
        date_end=date_end,
        build_url=lambda **kw: _build_url("financeiro_bp.contas_receber_cancelados", **kw),
    )


@financeiro_bp.route("/contas-receber/adicionar", methods=["POST"])
@login_required
def contas_receber_adicionar():
    cliente = (request.form.get("cliente") or "").strip()
    cnpj = (request.form.get("cnpj") or "").strip()
    contrato = (request.form.get("contrato") or "").strip()
    software = (request.form.get("software") or "").strip()
    data_primeira_pendencia = request.form.get("data_primeira_pendencia") or ""
    empresa_responsavel = (request.form.get("empresa_responsavel") or "").strip()
    valor = _parse_money(request.form.get("valor"))
    criado_por = _current_user_name()

    qt_pendencias = 1
    dias_atraso = 0
    total = valor

    try:
        db.session.execute(
            text(
                "INSERT INTO contas_receber "
                "(cliente, cnpj, contrato, software, data_primeira_pendencia, qt_pendencias, "
                "dias_atraso, total, empresa_responsavel, valor, criado_por, status) "
                "VALUES (:cliente, :cnpj, :contrato, :software, :data_primeira_pendencia, :qt_pendencias, "
                ":dias_atraso, :total, :empresa_responsavel, :valor, :criado_por, :status)"
            ),
            {
                "cliente": cliente,
                "cnpj": cnpj,
                "contrato": contrato,
                "software": software,
                "data_primeira_pendencia": data_primeira_pendencia,
                "qt_pendencias": qt_pendencias,
                "dias_atraso": dias_atraso,
                "total": total,
                "empresa_responsavel": empresa_responsavel,
                "valor": valor,
                "criado_por": criado_por,
                "status": "ABERTO",
            },
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao adicionar conta a receber: %s", exc)
        if _wants_json():
            return jsonify({"ok": False, "message": "Erro ao adicionar conta."}), 400
        flash("Erro ao adicionar conta.", "danger")
        return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))

    if _wants_json():
        return jsonify({"ok": True})
    flash("Conta adicionada com sucesso!", "success")
    return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))


@financeiro_bp.route("/contas-receber/subpendencia", methods=["POST"])
@login_required
def contas_receber_subpendencia():
    id_pai = _safe_int(request.form.get("id_pai"), 0)
    cliente = (request.form.get("cliente") or "").strip()
    cnpj = (request.form.get("cnpj") or "").strip()
    contrato = (request.form.get("contrato") or "").strip()
    software = (request.form.get("software") or "").strip()
    data_pendencia = request.form.get("data_pendencia") or ""
    valor = _parse_money(request.form.get("valor"))
    informacoes = (request.form.get("informacoes") or "").strip()

    data_primeira_pendencia = data_pendencia
    qt_pendencias = 1
    dias_atraso = 0

    try:
        db.session.execute(
            text(
                "INSERT INTO contas_receber "
                "(cliente, cnpj, contrato, software, data_primeira_pendencia, qt_pendencias, "
                "id_pai, dias_atraso, informacoes, valor, status) "
                "VALUES (:cliente, :cnpj, :contrato, :software, :data_primeira_pendencia, :qt_pendencias, "
                ":id_pai, :dias_atraso, :informacoes, :valor, :status)"
            ),
            {
                "cliente": cliente,
                "cnpj": cnpj,
                "contrato": contrato,
                "software": software,
                "data_primeira_pendencia": data_primeira_pendencia,
                "qt_pendencias": qt_pendencias,
                "id_pai": id_pai,
                "dias_atraso": dias_atraso,
                "informacoes": informacoes,
                "valor": valor,
                "status": "ABERTO",
            },
        )
        if id_pai:
            total_row = db.session.execute(
                text("SELECT total FROM contas_receber WHERE id = :id"),
                {"id": id_pai},
            ).fetchone()
            total_pai = (total_row[0] if total_row else 0) or 0
            total_atualizado = float(total_pai) + valor
            db.session.execute(
                text("UPDATE contas_receber SET total = :total WHERE id = :id"),
                {"total": total_atualizado, "id": id_pai},
            )
            db.session.execute(
                text("UPDATE contas_receber SET qt_pendencias = qt_pendencias + 1 WHERE id = :id"),
                {"id": id_pai},
            )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao adicionar subpendência: %s", exc)
        if _wants_json():
            return jsonify({"ok": False, "message": "Erro ao adicionar subpendência."}), 400
        flash("Erro ao adicionar subpendência.", "danger")
        return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))

    if _wants_json():
        return jsonify({"ok": True})
    flash("Subpendência adicionada.", "success")
    return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))


@financeiro_bp.route("/contas-receber/subpendencias", methods=["POST"])
@login_required
def contas_receber_subpendencias():
    id_pai = _safe_int(request.form.get("id_pai"), 0)
    if not id_pai:
        return jsonify({"ok": False, "items": []}), 400
    rows = db.session.execute(
        text("SELECT * FROM contas_receber WHERE id_pai = :id"),
        {"id": id_pai},
    ).fetchall()
    items = []
    for row in rows:
        item = _row_to_dict(row)
        item["data_primeira_pendencia_br"] = _format_date(item.get("data_primeira_pendencia"))
        item["valor_display"] = _format_money(item.get("valor"))
        items.append(item)
    return jsonify({"ok": True, "items": items, "count": len(items)})


@financeiro_bp.route("/contas-receber/editar", methods=["POST"])
@login_required
def contas_receber_editar():
    conta_id = _safe_int(request.form.get("id"), 0)
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400

    cliente = (request.form.get("cliente") or "").strip()
    cnpj = (request.form.get("cnpj") or "").strip()
    contrato = (request.form.get("contrato") or "").strip()
    empresa_responsavel = (request.form.get("empresa_responsavel") or "").strip()
    software = (request.form.get("software") or "").strip()
    data_primeira_pendencia = request.form.get("data_primeira_pendencia") or ""
    qt_pendencias = _safe_int(request.form.get("qt_pendencias"), 1)
    valor = _parse_money(request.form.get("valor"))

    try:
        soma_row = db.session.execute(
            text("SELECT SUM(valor) AS soma_pendencias FROM contas_receber WHERE id_pai = :id"),
            {"id": conta_id},
        ).fetchone()
        soma_pendencias = (soma_row[0] if soma_row else 0) or 0
        novo_total = float(valor) + float(soma_pendencias)

        db.session.execute(
            text(
                "UPDATE contas_receber SET cliente = :cliente, cnpj = :cnpj, contrato = :contrato, "
                "empresa_responsavel = :empresa_responsavel, software = :software, "
                "data_primeira_pendencia = :data_primeira_pendencia, qt_pendencias = :qt_pendencias, "
                "valor = :valor, total = :total WHERE id = :id"
            ),
            {
                "cliente": cliente,
                "cnpj": cnpj,
                "contrato": contrato,
                "empresa_responsavel": empresa_responsavel,
                "software": software,
                "data_primeira_pendencia": data_primeira_pendencia,
                "qt_pendencias": qt_pendencias,
                "valor": valor,
                "total": novo_total,
                "id": conta_id,
            },
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao editar conta: %s", exc)
        if _wants_json():
            return jsonify({"ok": False, "message": "Erro ao editar conta."}), 400
        flash("Erro ao editar conta.", "danger")
        return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))

    if _wants_json():
        return jsonify({"ok": True})
    flash("Conta atualizada.", "success")
    return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))


@financeiro_bp.route("/contas-receber/editar-sub", methods=["POST"])
@login_required
def contas_receber_editar_sub():
    conta_id = _safe_int(request.form.get("id"), 0)
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400

    cliente = (request.form.get("cliente") or "").strip()
    cnpj = (request.form.get("cnpj") or "").strip()
    contrato = (request.form.get("contrato") or "").strip()
    empresa_responsavel = (request.form.get("unidade") or "").strip()
    software = (request.form.get("software") or "").strip()
    data_primeira_pendencia = request.form.get("data_primeira_pendencia") or ""
    qt_pendencias = _safe_int(request.form.get("qt_pendencias"), 1)
    valor = _parse_money(request.form.get("valor"))

    try:
        parent_row = db.session.execute(
            text("SELECT id_pai FROM contas_receber WHERE id = :id"),
            {"id": conta_id},
        ).fetchone()
        id_pai = (parent_row[0] if parent_row else 0) or 0

        db.session.execute(
            text(
                "UPDATE contas_receber SET cliente = :cliente, cnpj = :cnpj, contrato = :contrato, "
                "empresa_responsavel = :empresa_responsavel, software = :software, "
                "data_primeira_pendencia = :data_primeira_pendencia, qt_pendencias = :qt_pendencias, "
                "valor = :valor WHERE id = :id"
            ),
            {
                "cliente": cliente,
                "cnpj": cnpj,
                "contrato": contrato,
                "empresa_responsavel": empresa_responsavel,
                "software": software,
                "data_primeira_pendencia": data_primeira_pendencia,
                "qt_pendencias": qt_pendencias,
                "valor": valor,
                "id": conta_id,
            },
        )

        if id_pai:
            valor_pai_row = db.session.execute(
                text("SELECT valor FROM contas_receber WHERE id = :id"),
                {"id": id_pai},
            ).fetchone()
            valor_pai = (valor_pai_row[0] if valor_pai_row else 0) or 0
            soma_row = db.session.execute(
                text("SELECT SUM(valor) AS soma_pendencias FROM contas_receber WHERE id_pai = :id"),
                {"id": id_pai},
            ).fetchone()
            soma_filhos = (soma_row[0] if soma_row else 0) or 0
            novo_total = float(valor_pai) + float(soma_filhos)
            db.session.execute(
                text("UPDATE contas_receber SET total = :total WHERE id = :id"),
                {"total": novo_total, "id": id_pai},
            )
        else:
            filhos = db.session.execute(
                text("SELECT id, valor FROM contas_receber WHERE id_pai = :id"),
                {"id": conta_id},
            ).fetchall()
            total_filhos = 0.0
            for filho in filhos:
                total_filhos += float(filho[1] or 0)
                db.session.execute(
                    text("UPDATE contas_receber SET total = :total WHERE id = :id"),
                    {"total": float(filho[1] or 0), "id": filho[0]},
                )
            novo_total = float(valor) + total_filhos
            db.session.execute(
                text("UPDATE contas_receber SET total = :total WHERE id = :id"),
                {"total": novo_total, "id": conta_id},
            )

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao editar subconta: %s", exc)
        if _wants_json():
            return jsonify({"ok": False, "message": "Erro ao editar subconta."}), 400
        flash("Erro ao editar subconta.", "danger")
        return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))

    if _wants_json():
        return jsonify({"ok": True})
    flash("Subconta atualizada.", "success")
    return redirect(request.referrer or url_for("financeiro_bp.contas_receber"))


@financeiro_bp.route("/contas-receber/quitar", methods=["POST"])
@login_required
def contas_receber_quitar():
    conta_id = _safe_int(request.form.get("id"), 0)
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400
    try:
        db.session.execute(
            text("UPDATE contas_receber SET status = 'QUITADO' WHERE id = :id"),
            {"id": conta_id},
        )
        db.session.execute(
            text("UPDATE contas_receber SET status = 'QUITADO' WHERE id_pai = :id"),
            {"id": conta_id},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao quitar conta: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao quitar conta."}), 400
    return jsonify({"ok": True})


@financeiro_bp.route("/contas-receber/update-info", methods=["POST"])
@login_required
def contas_receber_update_info():
    conta_id = _safe_int(request.form.get("id"), 0)
    new_info = (request.form.get("new_info") or "").strip()
    if not conta_id or not new_info:
        return jsonify({"ok": False, "message": "Dados incompletos."}), 400

    now = _now_br_string()
    user = _current_user_name()
    new_line = f'<li><span class="bullet"></span>{html.escape(new_info)} {now} - Usuário: {html.escape(user)}</li>'

    try:
        row = db.session.execute(
            text("SELECT informacoes FROM contas_receber WHERE id = :id"),
            {"id": conta_id},
        ).fetchone()
        current_info = row[0] if row else ""
        updated = f"{current_info}{new_line}" if current_info else new_line
        updated = f"<ul>{updated}</ul>"
        db.session.execute(
            text("UPDATE contas_receber SET informacoes = :info WHERE id = :id"),
            {"info": updated, "id": conta_id},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao atualizar informacoes: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao atualizar informacoes."}), 400
    return jsonify({"ok": True})


@financeiro_bp.route("/contas-receber/historico", methods=["POST"])
@login_required
def contas_receber_historico():
    conta_id = _safe_int(request.form.get("id"), 0)
    row = db.session.execute(
        text("SELECT informacoes FROM contas_receber WHERE id = :id"),
        {"id": conta_id},
    ).fetchone()
    historico = row[0] if row else ""
    if historico:
        historico = historico.replace("Usuário", "Usuário").replace("Usuários", "Usuários")
    return jsonify({"ok": True, "html": historico or ""})


@financeiro_bp.route("/contas-receber/valor", methods=["POST"])
@login_required
def contas_receber_valor():
    conta_id = _safe_int(request.form.get("id"), 0)
    valor_adicionado = _parse_money(request.form.get("valor"))
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400
    try:
        row = db.session.execute(
            text("SELECT valor, informacoes FROM contas_receber WHERE id = :id"),
            {"id": conta_id},
        ).fetchone()
        valor_atual = float(row[0] or 0) if row else 0.0
        novo_valor = valor_atual + valor_adicionado
        db.session.execute(
            text("UPDATE contas_receber SET valor = :valor WHERE id = :id"),
            {"valor": novo_valor, "id": conta_id},
        )
        data_atual = _now_br_string()
        user = _current_user_name()
        valor_anterior = f"R$ {_format_money(valor_atual)}"
        valor_add = f"R$ {_format_money(valor_adicionado)}"
        valor_total = f"R$ {_format_money(novo_valor)}"
        info_line = (
            f"Valor anterior: {valor_anterior} | Valor adicionado: {valor_add} | "
            f"Valor total: {valor_total} | Data: {data_atual} | Editado Por: {user}"
        )
        db.session.execute(
            text("UPDATE contas_receber SET informacoes = CONCAT(IFNULL(informacoes, ''), :info, '<br>') WHERE id = :id"),
            {"info": html.escape(info_line), "id": conta_id},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao atualizar valor: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao atualizar valor."}), 400
    return jsonify({"ok": True})


@financeiro_bp.route("/contas-receber/solicitar-bloqueio", methods=["POST"])
@login_required
def contas_receber_bloqueio():
    conta_id = _safe_int(request.form.get("id"), 0)
    data_bloqueio = request.form.get("data_bloqueio") or _now_br_string()
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400

    try:
        row = db.session.execute(
            text(
                "SELECT cliente, cnpj, contrato, software, data_primeira_pendencia, "
                "empresa_responsavel, valor, total, informacoes FROM contas_receber WHERE id = :id"
            ),
            {"id": conta_id},
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "Conta não encontrada."}), 404
        data_row = _row_to_dict(row)
        dias_atraso = _calc_dias_atraso(data_row.get("data_primeira_pendencia"))

        db.session.execute(
            text("UPDATE contas_receber SET dias_atraso = :dias, data_bloqueio = :data WHERE id = :id"),
            {"dias": dias_atraso, "data": data_bloqueio, "id": conta_id},
        )

        user = _current_user_name()
        new_info = f'<li><span class="bullet"></span>Usuário {html.escape(user)} solicitou o bloqueio em {data_bloqueio}</li>'
        current_info = data_row.get("informacoes") or ""
        updated = f"{current_info}{new_info}" if current_info else new_info
        db.session.execute(
            text("UPDATE contas_receber SET informacoes = :info WHERE id = :id"),
            {"info": updated, "id": conta_id},
        )
        db.session.commit()

        subject = f"Solicitação de bloqueio recebida - {data_row.get('cliente')}"
        message = (
            "Uma solicitação de bloqueio foi recebida para o seguinte contrato:<br>"
            f"Cliente: {data_row.get('cliente')}<br>"
            f"CNPJ: {data_row.get('cnpj')}<br>"
            f"Contrato: {data_row.get('contrato')}<br>"
            f"Software: {data_row.get('software')}<br>"
            f"Dias em atraso: {dias_atraso}<br>"
            f"Valor do Debito: R$ {_format_money(data_row.get('total'))}<br>"
            f"Empresa responsavel: {data_row.get('empresa_responsavel')}<br>"
            f"Bloqueio solicitado por: {user}<br>"
            f"Data de solicitação de bloqueio: {data_bloqueio}"
        )
        settings = _resolve_finance_smtp_settings()
        _dispatch_email_async(settings, subject, message)
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao solicitar bloqueio: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao solicitar bloqueio."}), 400

    return jsonify({"ok": True})


@financeiro_bp.route("/contas-receber/cancelamento", methods=["POST"])
@login_required
def contas_receber_cancelamento():
    conta_id = _safe_int(request.form.get("id"), 0)
    cancelamento_date = request.form.get("cancelamento") or _now_br_string()
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400

    try:
        db.session.execute(
            text("UPDATE contas_receber SET cancelamento = :data WHERE id = :id"),
            {"data": cancelamento_date, "id": conta_id},
        )
        row = db.session.execute(
            text(
                "SELECT cliente, cnpj, contrato, software, data_primeira_pendencia, "
                "empresa_responsavel, valor, total, informacoes FROM contas_receber WHERE id = :id"
            ),
            {"id": conta_id},
        ).fetchone()
        if not row:
            db.session.commit()
            return jsonify({"ok": True})
        data_row = _row_to_dict(row)
        dias_atraso = _calc_dias_atraso(data_row.get("data_primeira_pendencia"))
        user = _current_user_name()
        new_info = f'<li><span class="bullet"></span>Usuário {html.escape(user)} solicitou o Cancelamento em {cancelamento_date}</li>'
        current_info = data_row.get("informacoes") or ""
        updated = f"{current_info}{new_info}" if current_info else new_info
        db.session.execute(
            text("UPDATE contas_receber SET informacoes = :info WHERE id = :id"),
            {"info": updated, "id": conta_id},
        )
        db.session.commit()

        subject = f"Solicitação de cancelamento recebida - {data_row.get('cliente')}"
        message = (
            "Uma solicitação de cancelamento foi recebida para o seguinte contrato:<br>"
            f"Cliente: {data_row.get('cliente')}<br>"
            f"CNPJ: {data_row.get('cnpj')}<br>"
            f"Contrato: {data_row.get('contrato')}<br>"
            f"Software: {data_row.get('software')}<br>"
            f"Dias em atraso: {dias_atraso}<br>"
            f"Valor do Debito: R$ {_format_money(data_row.get('total'))}<br>"
            f"Empresa responsavel: {data_row.get('empresa_responsavel')}<br>"
            f"Cancelamento solicitado por: {user}<br>"
            f"Data de solicitação de cancelamento: {cancelamento_date}"
        )
        settings = _resolve_finance_smtp_settings()
        _dispatch_email_async(settings, subject, message)
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao solicitar cancelamento: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao solicitar cancelamento."}), 400

    return jsonify({"ok": True})


@financeiro_bp.route("/contas-receber/deferimento", methods=["POST"])
@login_required
def contas_receber_deferimento():
    conta_id = _safe_int(request.form.get("id"), 0)
    deferimento_date = request.form.get("deferimento") or _now_br_string()
    dias_atraso = _safe_int(request.form.get("diasAtraso"), 0)
    if not conta_id:
        return jsonify({"ok": False, "message": "Conta inválida."}), 400

    try:
        db.session.execute(
            text(
                "UPDATE contas_receber SET deferimento_cancelamento = :data, dias_atraso = :dias, "
                "status = :status WHERE id = :id"
            ),
            {"data": deferimento_date, "dias": dias_atraso, "status": "FECHADO", "id": conta_id},
        )
        db.session.commit()

        subject = "Cancelamento deferido"
        message = f"O cancelamento foi deferido em: {deferimento_date}"
        settings = _resolve_finance_smtp_settings()
        _dispatch_email_async(settings, subject, message)
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao deferir cancelamento: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao deferir cancelamento."}), 400

    return jsonify({"ok": True})


@financeiro_bp.route("/contas-receber/excluir", methods=["POST"])
@login_required
def contas_receber_excluir():
    conta_id = _safe_int(request.form.get("id"), 0)
    if not conta_id:
        return jsonify({"ok": False, "message": "Registro inválido."}), 400
    try:
        exists = db.session.execute(
            text("SELECT 1 FROM contas_receber WHERE id = :id"),
            {"id": conta_id},
        ).fetchone()
        if not exists:
            db.session.rollback()
            current_app.logger.warning("Tentativa de excluir pendência inexistente: %s", conta_id)
            return jsonify({"ok": False, "message": "Registro não localizado."}), 404
        result = db.session.execute(
            text(
                "DELETE FROM contas_receber "
                "WHERE id = :id OR id_pai = :id"
            ),
            {"id": conta_id},
        )
        current_app.logger.info("Excluiu pendência %s (subexistência %s rows)", conta_id, result.rowcount)
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Erro ao excluir pendência.")
        return jsonify({"ok": False, "message": "Não foi possível excluir o registro."}), 400
    return jsonify({"ok": True})


@financeiro_bp.route("/verifica-empresa")
@login_required
def verifica_empresa():
    cnpj = (request.args.get("cnpj") or "").strip()
    if not cnpj:
        return ""
    cnpj = re.sub(r"\D+", "", cnpj)

    row = db.session.execute(
        text("SELECT cliente FROM empresa WHERE cnpj = :cnpj"),
        {"cnpj": cnpj},
    ).fetchone()
    if row and row[0]:
        return str(row[0])

    try:
        resp = requests.get(f"https://www.receitaws.com.br/v1/cnpj/{cnpj}", timeout=12)
        if resp.ok:
            data = resp.json()
            if data.get("nome"):
                return str(data.get("nome"))
    except Exception:
        current_app.logger.warning("Falha ao consultar receitaws para CNPJ %s", cnpj)

    return "CNPJ não encontrado."



def _parse_filtro_mes(filtro: str | None) -> tuple[date, str, int]:
    today = date.today()
    if filtro:
        try:
            mes, ano = filtro.split("-")
            dt = date(int(ano), int(mes), 1)
            return dt, mes.zfill(2), int(ano)
        except Exception:
            pass
    return date(today.year, today.month, 1), f"{today.month:02d}", today.year


def _fetch_cota_mensal(filter_date: date) -> list[dict[str, Any]]:
    sql = text(
        "SELECT * FROM cota_mensal "
        "WHERE Data_Inicio <= LAST_DAY(:filter_date) AND Data_Fim >= :filter_date"
    )
    rows = db.session.execute(sql, {"filter_date": filter_date}).fetchall()
    items = []
    for row in rows:
        item = _row_to_dict(row)
        item["Data_Inicio_br"] = _format_date(item.get("Data_Inicio"))
        item["Data_Fim_br"] = _format_date(item.get("Data_Fim"))
        item["Valor_Arrecadado_display"] = _format_money(item.get("Valor_Arrecadado"))
        item["Valor_Pago_display"] = _format_money(item.get("Valor_Pago"))
        item["Valor_Mes_display"] = _format_money(item.get("Valor_Mes"))
        items.append(item)
    return items


def _fetch_cota_totais(filter_date: date) -> list[dict[str, Any]]:
    sql = text(
        "SELECT Unidade_Responsavel, SUM(Valor_Arrecadado) AS Total_Arrecadado, "
        "SUM(Valor_Pago) AS Total_Pago FROM cota_mensal "
        "WHERE Data_Inicio <= LAST_DAY(:filter_date) AND Data_Fim >= :filter_date "
        "GROUP BY Unidade_Responsavel"
    )
    rows = db.session.execute(sql, {"filter_date": filter_date}).fetchall()
    items = []
    for row in rows:
        item = _row_to_dict(row)
        item["Total_Arrecadado_display"] = _format_money(item.get("Total_Arrecadado"))
        item["Total_Pago_display"] = _format_money(item.get("Total_Pago"))
        items.append(item)
    return items


def _fetch_periodo_trimestre_atual() -> dict[str, Any] | None:
    row = db.session.execute(
        text(
            "SELECT * FROM periodo_trimestre WHERE data_fim >= CURDATE() "
            "ORDER BY data_inicial ASC LIMIT 1"
        )
    ).fetchone()
    return _row_to_dict(row) if row else None


@financeiro_bp.route("/cota")
@login_required
def cota_dashboard():
    filtro = request.args.get("filtro")
    filter_date, mes_str, ano_int = _parse_filtro_mes(filtro)

    cota_mensal = _fetch_cota_mensal(filter_date)
    totais = _fetch_cota_totais(filter_date)

    meta_valor = float(cota_mensal[0].get("Valor_Mes") or 0) if cota_mensal else 0.0
    valor_arrecadado_total = sum(float(item.get("Valor_Arrecadado") or 0) for item in cota_mensal)
    valor_pago_total = sum(float(item.get("Valor_Pago") or 0) for item in cota_mensal)

    chart_mensal = [
        [item.get("Unidade_Responsavel"), float(item.get("Total_Arrecadado") or 0)]
        for item in totais
    ]

    periodo = _fetch_periodo_trimestre_atual()
    cota_trimestral = []
    chart_trimestral = []
    trimestre_totais = {"faturado": 0.0, "recebido": 0.0}
    if periodo:
        rows = db.session.execute(
            text(
                "SELECT * FROM cota_trimestral WHERE Data_inicio >= :inicio AND Data_fim <= :fim"
            ),
            {"inicio": periodo.get("data_inicial"), "fim": periodo.get("data_fim")},
        ).fetchall()
        for row in rows:
            item = _row_to_dict(row)
            item["Data_inicio_br"] = _format_date(item.get("Data_inicio"))
            item["Data_fim_br"] = _format_date(item.get("Data_fim"))
            item["Valor_faturado_display"] = _format_money(item.get("Valor_faturado"))
            item["Valor_recebido_display"] = _format_money(item.get("Valor_recebido"))
            cota_trimestral.append(item)
        summary_rows = db.session.execute(
            text(
                "SELECT Nome_mes, SUM(Valor_faturado) AS Valor_faturado, "
                "SUM(Valor_recebido) AS Valor_recebido FROM cota_trimestral "
                "WHERE Data_inicio >= :inicio AND Data_fim <= :fim GROUP BY Nome_mes"
            ),
            {"inicio": periodo.get("data_inicial"), "fim": periodo.get("data_fim")},
        ).fetchall()
        for row in summary_rows:
            item = _row_to_dict(row)
            nome = item.get("Nome_mes")
            faturado = float(item.get("Valor_faturado") or 0)
            recebido = float(item.get("Valor_recebido") or 0)
            chart_trimestral.append([nome, faturado, recebido])
            trimestre_totais["faturado"] += faturado
            trimestre_totais["recebido"] += recebido

    return render_template(
        "admin/financeiro/cota.html",
        filtro=filtro,
        mes_selecionado=mes_str,
        ano_selecionado=ano_int,
        cota_mensal=cota_mensal,
        totais=totais,
        meta_valor=meta_valor,
        meta_valor_display=_format_money(meta_valor),
        valor_arrecadado_total=valor_arrecadado_total,
        valor_arrecadado_total_display=_format_money(valor_arrecadado_total),
        valor_pago_total=valor_pago_total,
        valor_pago_total_display=_format_money(valor_pago_total),
        chart_mensal=chart_mensal,
        periodo_trimestre=periodo,
        cota_trimestral=cota_trimestral,
        chart_trimestral=chart_trimestral,
        trimestre_totais=trimestre_totais,
        trimestre_totais_display={
            "faturado": _format_money(trimestre_totais["faturado"]),
            "recebido": _format_money(trimestre_totais["recebido"]),
        },
    )


@financeiro_bp.route("/cota/adicionar", methods=["POST"])
@login_required
def cota_adicionar():
    valor_mes = _parse_money(request.form.get("valorMes"))
    data_inicio_raw = request.form.get("dataInicio")
    if not data_inicio_raw:
        flash("Informe a data de inicio.", "warning")
        return redirect(url_for("financeiro_bp.cota_dashboard"))

    start = _parse_date(data_inicio_raw)
    if not start:
        flash("Data inválida.", "warning")
        return redirect(url_for("financeiro_bp.cota_dashboard"))

    last_day = date(start.year, start.month, calendar.monthrange(start.year, start.month)[1])
    usuario = _current_user_name()

    try:
        for unidade in FINANCEIRO_UNIDADES_COTA:
            db.session.execute(
                text(
                    "INSERT INTO cota_mensal (Valor_Mes, Valor_Arrecadado, Valor_Pago, Data_Inicio, "
                    "Data_Fim, Usuario_Adicionou, Unidade_Responsavel) "
                    "VALUES (:valor_mes, 0, 0, :data_inicio, :data_fim, :usuario, :unidade)"
                ),
                {
                    "valor_mes": valor_mes,
                    "data_inicio": start,
                    "data_fim": last_day,
                    "usuario": usuario,
                    "unidade": unidade,
                },
            )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao adicionar cota mensal: %s", exc)
        flash("Erro ao adicionar cota mensal.", "danger")
        return redirect(url_for("financeiro_bp.cota_dashboard"))

    flash("Cota mensal adicionada.", "success")
    return redirect(url_for("financeiro_bp.cota_dashboard"))


@financeiro_bp.route("/cota/valor", methods=["POST"])
@login_required
def cota_valor():
    cota_id = _safe_int(request.form.get("id"), 0)
    valor = _parse_money(request.form.get("valor"))
    tipo = request.form.get("tipo") or ""
    if not cota_id:
        return jsonify({"ok": False, "message": "Registro inválido."}), 400

    try:
        row = db.session.execute(
            text("SELECT Valor_Arrecadado FROM cota_mensal WHERE ID = :id"),
            {"id": cota_id},
        ).fetchone()
        valor_atual = float(row[0] or 0) if row else 0.0
        novo_valor = valor_atual + valor if tipo == "adicionar" else valor_atual - valor
        db.session.execute(
            text("UPDATE cota_mensal SET Valor_Arrecadado = :valor WHERE ID = :id"),
            {"valor": novo_valor, "id": cota_id},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao atualizar valor arrecadado: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao atualizar valor."}), 400

    return jsonify({"ok": True, "valor": _format_money(novo_valor)})


@financeiro_bp.route("/cota/valor-pago", methods=["POST"])
@login_required
def cota_valor_pago():
    cota_id = _safe_int(request.form.get("id"), 0)
    valor = _parse_money(request.form.get("valor"))
    tipo = request.form.get("tipo") or ""
    if not cota_id:
        return jsonify({"ok": False, "message": "Registro inválido."}), 400

    try:
        row = db.session.execute(
            text("SELECT Valor_Pago FROM cota_mensal WHERE ID = :id"),
            {"id": cota_id},
        ).fetchone()
        valor_atual = float(row[0] or 0) if row else 0.0
        novo_valor = valor_atual + valor if tipo == "adicionar" else valor_atual - valor
        novo_valor = abs(novo_valor)
        db.session.execute(
            text("UPDATE cota_mensal SET Valor_Pago = :valor WHERE ID = :id"),
            {"valor": novo_valor, "id": cota_id},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao atualizar valor pago: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao atualizar valor pago."}), 400

    return jsonify({"ok": True, "valor": _format_money(novo_valor)})


@financeiro_bp.route("/cota/fechar-mes", methods=["POST"])
@login_required
def cota_fechar_mes():
    filtro = (request.form.get("filtro") or "").strip()
    if not filtro:
        return jsonify({"ok": False, "message": "Informe o mês."}), 400
    try:
        mes, ano = filtro.split("-")
        data_mes = date(int(ano), int(mes), 1)
    except Exception:
        return jsonify({"ok": False, "message": "Mês inválido."}), 400

    try:
        db.session.execute(
            text(
                "UPDATE cota_mensal SET status = 'FECHADO', data_status = CURDATE() "
                "WHERE DATE_FORMAT(Data_Inicio, '%Y-%m') = :mes AND status = 'ABERTO'"
            ),
            {"mes": data_mes.strftime("%Y-%m")},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao fechar mes da cota: %s", exc)
        return jsonify({"ok": False, "message": "Erro ao fechar mês."}), 400
    return jsonify({"ok": True})


@financeiro_bp.route("/cota/trimestre", methods=["POST"])
@login_required
def cota_trimestre():
    data_inicio_raw = request.form.get("dataInicio")
    if not data_inicio_raw:
        flash("Informe a data de inicio.", "warning")
        return redirect(url_for("financeiro_bp.cota_dashboard"))

    data_inicio = _parse_date(data_inicio_raw)
    if not data_inicio:
        flash("Data inválida.", "warning")
        return redirect(url_for("financeiro_bp.cota_dashboard"))

    try:
        rows = db.session.execute(
            text("SELECT * FROM cota_mensal WHERE Data_Inicio = :data"),
            {"data": data_inicio},
        ).fetchall()
        if not rows:
            flash("Nenhuma cota mensal encontrada para o mes.", "warning")
            return redirect(url_for("financeiro_bp.cota_dashboard"))

        ids = []
        data_inicio_trimestre = None
        data_fim_trimestre = None
        valor_arrecadado_total = 0.0
        valor_pago_total = 0.0

        for row in rows:
            item = _row_to_dict(row)
            ids.append(str(item.get("ID")))
            data_inicio_item = _parse_date(item.get("Data_Inicio"))
            data_fim_item = _parse_date(item.get("Data_Fim"))
            if data_inicio_item and (data_inicio_trimestre is None or data_inicio_item < data_inicio_trimestre):
                data_inicio_trimestre = data_inicio_item
            if data_fim_item and (data_fim_trimestre is None or data_fim_item > data_fim_trimestre):
                data_fim_trimestre = data_fim_item
            valor_arrecadado_total += float(item.get("Valor_Arrecadado") or 0)
            valor_pago_total += float(item.get("Valor_Pago") or 0)

        nome_mes = MESES_PT.get(data_inicio.strftime("%B"), data_inicio.strftime("%B"))
        usuario = _current_user_name()

        result = db.session.execute(
            text(
                "INSERT INTO cota_trimestral "
                "(ID_cota_mensal, Nome_mes, Valor_faturado, Valor_recebido, Data_inicio, Data_fim, Quem_adicionou) "
                "VALUES (:ids, :nome_mes, :valor_faturado, :valor_recebido, :data_inicio, :data_fim, :usuario)"
            ),
            {
                "ids": ",".join(ids),
                "nome_mes": nome_mes,
                "valor_faturado": valor_arrecadado_total,
                "valor_recebido": valor_pago_total,
                "data_inicio": data_inicio_trimestre,
                "data_fim": data_fim_trimestre,
                "usuario": usuario,
            },
        )
        id_trimestre = getattr(result, "lastrowid", None)

        data_fim_periodo = date(
            data_inicio_trimestre.year,
            data_inicio_trimestre.month,
            calendar.monthrange(data_inicio_trimestre.year, data_inicio_trimestre.month)[1],
        )
        data_fim_periodo = data_fim_periodo.replace(
            month=((data_inicio_trimestre.month - 1 + 2) % 12) + 1,
            year=data_inicio_trimestre.year + ((data_inicio_trimestre.month - 1 + 2) // 12),
        )
        data_fim_periodo = date(
            data_fim_periodo.year,
            data_fim_periodo.month,
            calendar.monthrange(data_fim_periodo.year, data_fim_periodo.month)[1],
        )

        existe = db.session.execute(
            text(
                "SELECT * FROM periodo_trimestre WHERE :data_inicio BETWEEN Mes1 AND Mes3"
            ),
            {"data_inicio": data_inicio_trimestre},
        ).fetchone()

        if not existe:
            mes2 = date(
                data_inicio_trimestre.year + (data_inicio_trimestre.month // 12),
                ((data_inicio_trimestre.month - 1 + 1) % 12) + 1,
                data_inicio_trimestre.day,
            )
            mes3 = date(
                data_inicio_trimestre.year + ((data_inicio_trimestre.month + 1) // 12),
                ((data_inicio_trimestre.month - 1 + 2) % 12) + 1,
                data_inicio_trimestre.day,
            )
            db.session.execute(
                text(
                    "INSERT INTO periodo_trimestre (id_trimestre, data_inicial, data_fim, Mes1, Mes2, Mes3) "
                    "VALUES (:id, :data_inicial, :data_fim, :mes1, :mes2, :mes3)"
                ),
                {
                    "id": id_trimestre,
                    "data_inicial": data_inicio_trimestre,
                    "data_fim": data_fim_periodo,
                    "mes1": data_inicio_trimestre,
                    "mes2": mes2,
                    "mes3": mes3,
                },
            )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erro ao criar cota trimestral: %s", exc)
        flash("Erro ao criar cota trimestral.", "danger")
        return redirect(url_for("financeiro_bp.cota_dashboard"))

    flash("Cota trimestral criada.", "success")
    return redirect(url_for("financeiro_bp.cota_dashboard"))


@financeiro_bp.route("/cota/pdf-mensal")
@login_required
def cota_pdf_mensal():
    mes = _safe_int(request.args.get("mes"), 0)
    ano = _safe_int(request.args.get("ano"), 0)
    today = date.today()
    if not mes or not ano:
        mes = today.month
        ano = today.year
    filtro = date(ano, mes, 1)
    totais = _fetch_cota_totais(filtro)
    context = {
        "mes": mes,
        "ano": ano,
        "rows": totais,
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    pdf_bytes = render_proposta_html_pdf("financeiro/cota_pdf_mensal.html", context)
    return send_file(
        io.BytesIO(pdf_bytes),
        download_name="totais_cota_mensal.pdf",
        as_attachment=False,
    )


@financeiro_bp.route("/cota/pdf-trimestre")
@login_required
def cota_pdf_trimestre():
    data_inicial_raw = request.args.get("data_inicial")
    data_inicial = _parse_date(data_inicial_raw) or date.today()

    periodo = db.session.execute(
        text("SELECT * FROM periodo_trimestre WHERE data_inicial = :data"),
        {"data": data_inicial},
    ).fetchone()
    if not periodo:
        flash("Trimestre não encontrado.", "warning")
        return redirect(url_for("financeiro_bp.cota_dashboard"))
    periodo = _row_to_dict(periodo)
    meses = [periodo.get("Mes1"), periodo.get("Mes2"), periodo.get("Mes3")]
    rows = db.session.execute(
        text("SELECT * FROM cota_trimestral WHERE Data_inicio IN :meses").bindparams(
            bindparam("meses", expanding=True)
        ),
        {"meses": meses},
    ).fetchall()
    items = []
    total_faturado = 0.0
    total_recebido = 0.0
    for row in rows:
        item = _row_to_dict(row)
        faturado = float(item.get("Valor_faturado") or 0)
        recebido = float(item.get("Valor_recebido") or 0)
        total_faturado += faturado
        total_recebido += recebido
        item["Valor_faturado_display"] = _format_money(faturado)
        item["Valor_recebido_display"] = _format_money(recebido)
        inicio_data = _parse_date(item.get("Data_inicio"))
        if inicio_data:
            item["Nome_mes"] = MESES_PT.get(inicio_data.strftime("%B"), inicio_data.strftime("%B"))
        items.append(item)

    context = {
        "rows": items,
        "total_faturado": _format_money(total_faturado),
        "total_recebido": _format_money(total_recebido),
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    pdf_bytes = render_proposta_html_pdf("financeiro/cota_pdf_trimestre.html", context)
    return send_file(
        io.BytesIO(pdf_bytes),
        download_name="cota_trimestral.pdf",
        as_attachment=False,
    )


@financeiro_bp.route("/cota/enviar-email", methods=["POST"])
@login_required
def cota_enviar_email():
    tabela = request.form.get("tabela") or ""
    if not tabela:
        return jsonify({"ok": False, "message": "Tabela vazia."}), 400
    body = f"Resumo do faturamento:<br><br>{tabela}"
    ok = _send_email(_resolve_cota_smtp_settings(), "Cota - Faturamento ate o momento", body)
    return jsonify({"ok": ok})


@financeiro_bp.route("/contas-receber/export")
@login_required
def contas_receber_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    search = (request.args.get("search") or "").strip()
    date_start = (request.args.get("date_start") or "").strip()
    date_end = (request.args.get("date_end") or "").strip()
    page_mode = request.args.get("page_mode") or "aberto"

    if page_mode == "cancelados":
        statuses = ["FECHADO", "QUITADO"]
    else:
        statuses = ["ABERTO"]

    # Fetch ALL records matching filters
    contas, _ = _fetch_contas(
        statuses=statuses,
        search=search,
        date_start=date_start,
        date_end=date_end,
        page=1,
        per_page=1000000,
        order="DESC" if page_mode == "cancelados" else "ASC",
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contas a Receber"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B3B8C", end_color="0B3B8C", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "ID", "Cliente", "CNPJ", "Contrato", "Software", 
        "Data Primeira Pendência", "Valor", "Dias em Atraso",
        "Empresa Responsável", "Criado Por", "Status"
    ]
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    for item in contas:
        data_pendencia = item.get("data_primeira_pendencia")
        if data_pendencia:
            try:
                dt = _parse_date(data_pendencia)
                data_str = dt.strftime("%d/%m/%Y") if dt else str(data_pendencia)
            except Exception:
                data_str = str(data_pendencia)
        else:
            data_str = ""

        row_data = [
            item.get("id"),
            item.get("cliente"),
            item.get("cnpj"),
            item.get("contrato"),
            item.get("software"),
            data_str,
            item.get("total") or item.get("valor") or 0.0,
            item.get("dias_atraso_calc") or item.get("dias_atraso") or 0,
            item.get("empresa_responsavel"),
            item.get("criado_por"),
            (item.get("status") or "").upper()
        ]
        ws.append(row_data)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            if col_idx in (1, 6, 8, 11):
                cell.alignment = align_center
            elif col_idx == 7:
                cell.number_format = 'R$#,##0.00'
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"contas_receber_{page_mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



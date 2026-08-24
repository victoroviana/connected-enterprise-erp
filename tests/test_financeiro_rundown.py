from __future__ import annotations

import io
import pytest
from datetime import date, datetime
from unittest.mock import patch, MagicMock

import openpyxl
from sqlalchemy import event, text
from sqlalchemy.engine import Engine

from platform_app import create_app
from extensions import db
from modules.propostas.models import User
from modules.financeiro.blueprints.financeiro import (
    _safe_int,
    _parse_money,
    _format_money,
    _parse_date,
    _calc_dias_atraso,
    _sanitize_mojibake,
    _parse_filtro_mes,
    FINANCEIRO_UNIDADES_COTA,
)


class TestConfig:
    TESTING = True
    WTF_CSRF_ENABLED = False
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    MAIL_ENABLED = False
    MAIL_SERVER = "smtp.test.com"
    MAIL_DEFAULT_SENDER = "financeiro@sollustecnologia.com"
    MAIL_SENDER = "financeiro@sollustecnologia.com"
    SECRET_KEY = "test-secret-financeiro-key"


@event.listens_for(Engine, "connect")
def _setup_sqlite_functions(dbapi_connection, connection_record):
    """Register custom SQLite functions to emulate MySQL functions used in financeiro queries."""
    if type(dbapi_connection).__module__.startswith("sqlite"):
        def last_day(val):
            if not val:
                return None
            try:
                val_str = str(val)[:10]
                dt = datetime.strptime(val_str, "%Y-%m-%d").date()
                import calendar
                _, last = calendar.monthrange(dt.year, dt.month)
                return date(dt.year, dt.month, last).strftime("%Y-%m-%d")
            except Exception:
                return str(val)

        def curdate():
            return date.today().strftime("%Y-%m-%d")

        def date_format(val, fmt):
            if not val:
                return ""
            try:
                val_str = str(val)[:10]
                dt = datetime.strptime(val_str, "%Y-%m-%d")
                return dt.strftime("%Y-%m") if fmt == "%Y-%m" else dt.strftime("%Y-%m-%d")
            except Exception:
                return str(val)

        def concat(*args):
            return "".join([str(a) for a in args if a is not None])

        def ifnull(val, default):
            return default if val is None else val

        dbapi_connection.create_function("LAST_DAY", 1, last_day)
        dbapi_connection.create_function("CURDATE", 0, curdate)
        dbapi_connection.create_function("DATE_FORMAT", 2, date_format)
        dbapi_connection.create_function("CONCAT", -1, concat)
        dbapi_connection.create_function("IFNULL", 2, ifnull)


@pytest.fixture
def app():
    app_instance = create_app(TestConfig)
    with app_instance.app_context():
        db.create_all()

        # Create finance legacy tables
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS contas_receber (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente VARCHAR(255),
                cnpj VARCHAR(50),
                contrato VARCHAR(100),
                software VARCHAR(100),
                data_primeira_pendencia VARCHAR(50),
                qt_pendencias INTEGER DEFAULT 1,
                dias_atraso INTEGER DEFAULT 0,
                total FLOAT DEFAULT 0.0,
                empresa_responsavel VARCHAR(255),
                valor FLOAT DEFAULT 0.0,
                criado_por VARCHAR(100),
                status VARCHAR(50) DEFAULT 'ABERTO',
                id_pai INTEGER DEFAULT 0,
                data_bloqueio VARCHAR(50),
                cancelamento VARCHAR(50),
                deferimento_cancelamento VARCHAR(50),
                informacoes TEXT
            )
        """))

        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS cota_mensal (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                Valor_Mes FLOAT DEFAULT 0.0,
                Valor_Arrecadado FLOAT DEFAULT 0.0,
                Valor_Pago FLOAT DEFAULT 0.0,
                Data_Inicio DATE,
                Data_Fim DATE,
                Usuario_Adicionou VARCHAR(100),
                Unidade_Responsavel VARCHAR(100),
                status VARCHAR(50) DEFAULT 'ABERTO',
                data_status DATE
            )
        """))

        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS cota_trimestral (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_cota_mensal TEXT,
                Nome_mes VARCHAR(50),
                Valor_faturado FLOAT DEFAULT 0.0,
                Valor_recebido FLOAT DEFAULT 0.0,
                Data_inicio DATE,
                Data_fim DATE,
                Quem_adicionou VARCHAR(100)
            )
        """))

        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS periodo_trimestre (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_trimestre INTEGER,
                data_inicial DATE,
                data_fim DATE,
                Mes1 DATE,
                Mes2 DATE,
                Mes3 DATE
            )
        """))

        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS empresa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente VARCHAR(255),
                cnpj VARCHAR(50)
            )
        """))

        # Create admin user
        admin = User(
            usuario="admin_fin",
            nome_completo="Admin Financeiro",
            email="admin_fin@sollus.com",
            password_hash="pbkdf2:sha256:dummy",
            tipo="admin",
            role="admin",
            is_active=True,
        )
        # Create regular user without finance perms
        regular = User(
            usuario="regular_user",
            nome_completo="Usuario Comum",
            email="regular@sollus.com",
            password_hash="pbkdf2:sha256:dummy",
            tipo="user",
            role="user",
            is_active=True,
            permissions={},
        )
        # Create finance user with department
        finance_user = User(
            usuario="finance_dept_user",
            nome_completo="Usuario Financeiro",
            email="financeiro@sollus.com",
            password_hash="pbkdf2:sha256:dummy",
            tipo="user",
            role="user",
            is_active=True,
            permissions={"financeiro": True},
        )

        db.session.add_all([admin, regular, finance_user])
        db.session.commit()

        yield app_instance


@pytest.fixture
def client(app):
    return app.test_client()


def login(client, username="admin_fin"):
    with client.session_transaction() as sess:
        user = User.query.filter_by(usuario=username).first()
        sess["_user_id"] = str(user.id)
        sess["logged_in_user"] = user.nome_completo
        sess["tipo"] = user.tipo


# ==============================================================================
# 1. UNIT TESTS FOR HELPER FUNCTIONS
# ==============================================================================

def test_helper_safe_int():
    assert _safe_int("10") == 10
    assert _safe_int("abc", default=5) == 5
    assert _safe_int(None, default=1) == 1


def test_helper_parse_money():
    assert _parse_money("R$ 1.500,50") == 1500.50
    assert _parse_money("1.234,56") == 1234.56
    assert _parse_money(None) == 0.0
    assert _parse_money("") == 0.0
    assert _parse_money("invalido") == 0.0


def test_helper_format_money():
    assert _format_money(1500.5) == "1.500,50"
    assert _format_money(0) == "0,00"
    assert _format_money("invalid") == "0,00"


def test_helper_parse_date():
    d = date(2026, 7, 28)
    assert _parse_date(d) == d
    assert _parse_date("2026-07-28") == d
    assert _parse_date("28/07/2026") == d
    assert _parse_date(None) is None
    assert _parse_date("invalid-date") is None


def test_helper_calc_dias_atraso():
    past_date = "2020-01-01"
    assert _calc_dias_atraso(past_date) > 1000
    assert _calc_dias_atraso("2099-01-01") == 0
    assert _calc_dias_atraso(None) == 0


def test_helper_sanitize_mojibake():
    assert _sanitize_mojibake("teste") == "teste"
    assert _sanitize_mojibake(None) is None
    assert _sanitize_mojibake(["item1", "item2"]) == ["item1", "item2"]


def test_helper_parse_filtro_mes():
    dt, mes, ano = _parse_filtro_mes("05-2025")
    assert dt == date(2025, 5, 1)
    assert mes == "05"
    assert ano == 2025

    today = date.today()
    dt_def, mes_def, ano_def = _parse_filtro_mes(None)
    assert dt_def == date(today.year, today.month, 1)
    assert mes_def == f"{today.month:02d}"
    assert ano_def == today.year


# ==============================================================================
# 2. PERMISSIONS & ACCESS CONTROL TESTS
# ==============================================================================

def test_permissions_unauthorized(client):
    # Without login -> Flask-Login redirects to login
    response = client.get("/financeiro/contas-receber")
    assert response.status_code in (302, 401)


def test_permissions_forbidden_user(client):
    login(client, username="regular_user")
    response = client.get("/financeiro/contas-receber")
    assert response.status_code == 302
    assert "/financeiro/sem-permissao" in response.headers["Location"]

    # JSON request gets 403
    response_json = client.get("/financeiro/contas-receber", headers={"X-Requested-With": "XMLHttpRequest"})
    assert response_json.status_code == 403
    data = response_json.get_json()
    assert data["ok"] is False


def test_permissions_admin_allowed(client):
    login(client, username="admin_fin")
    response = client.get("/financeiro/contas-receber")
    assert response.status_code == 200


def test_permissions_finance_user_allowed(client):
    login(client, username="finance_dept_user")
    response = client.get("/financeiro/contas-receber")
    assert response.status_code == 200


# ==============================================================================
# 3. CONTAS A RECEBER ENDPOINTS (CRUD & WORKFLOWS)
# ==============================================================================

def test_contas_receber_list(client):
    login(client, username="admin_fin")
    response = client.get("/financeiro/contas-receber")
    assert response.status_code == 200
    assert b"contas_receber" in response.data or b"Pend\xc3\xaanclas" in response.data or b"Pend" in response.data


def test_contas_receber_adicionar(client):
    login(client, username="admin_fin")

    # Form submission
    payload = {
        "cliente": "Empresa Teste LTDA",
        "cnpj": "12.345.678/0001-90",
        "contrato": "CTR-1001",
        "software": "Sollus Ponto",
        "data_primeira_pendencia": "2026-07-01",
        "empresa_responsavel": "SOLLUS RJ",
        "valor": "R$ 1.500,00",
    }
    response = client.post(
        "/financeiro/contas-receber/adicionar",
        data=payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data["ok"] is True

    # Check DB row
    row = db.session.execute(
        text("SELECT * FROM contas_receber WHERE cliente = 'Empresa Teste LTDA'")
    ).fetchone()
    assert row is not None
    assert row.cliente == "Empresa Teste LTDA"
    assert row.valor == 1500.0
    assert row.total == 1500.0
    assert row.status == "ABERTO"


def test_contas_receber_subpendencia(client):
    login(client, username="admin_fin")

    # First add parent account
    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, cnpj, contrato, valor, total, status, id_pai)
        VALUES ('Cliente Pai', '11.111.111/0001-11', 'CTR-PAI', 1000.0, 1000.0, 'ABERTO', 0)
    """))
    db.session.commit()
    pai = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Pai'")).fetchone()
    pai_id = pai[0]

    # Add subpendencia
    sub_payload = {
        "id_pai": str(pai_id),
        "cliente": "Cliente Pai",
        "cnpj": "11.111.111/0001-11",
        "contrato": "CTR-PAI",
        "software": "Modulo Adicional",
        "data_pendencia": "2026-07-15",
        "valor": "500,00",
        "informacoes": "Subpendencia de servico extra",
    }
    res = client.post(
        "/financeiro/contas-receber/subpendencia",
        data=sub_payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    # Check updated total of parent
    pai_row = db.session.execute(text("SELECT total, qt_pendencias FROM contas_receber WHERE id = :id"), {"id": pai_id}).fetchone()
    assert pai_row.total == 1500.0
    assert pai_row.qt_pendencias == 2

    # Fetch subpendencias JSON list
    list_res = client.post(
        "/financeiro/contas-receber/subpendencias",
        data={"id_pai": str(pai_id)},
    )
    assert list_res.status_code == 200
    sub_data = list_res.get_json()
    assert sub_data["ok"] is True
    assert len(sub_data["items"]) == 1
    assert sub_data["items"][0]["valor_display"] == "500,00"


def test_contas_receber_editar_and_editar_sub(client):
    login(client, username="admin_fin")

    # Insert parent
    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, cnpj, contrato, software, valor, total, status, id_pai)
        VALUES ('Cliente Edit', '22.222.222/0001-22', 'CTR-EDIT', 'Old Soft', 800.0, 800.0, 'ABERTO', 0)
    """))
    db.session.commit()
    row = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Edit'")).fetchone()
    cid = row[0]

    # Edit parent
    edit_payload = {
        "id": str(cid),
        "cliente": "Cliente Edit Atualizado",
        "cnpj": "22.222.222/0001-22",
        "contrato": "CTR-EDIT-NEW",
        "software": "New Soft",
        "empresa_responsavel": "SOLLUS SP",
        "data_primeira_pendencia": "2026-07-05",
        "qt_pendencias": "1",
        "valor": "950,00",
    }
    res = client.post("/financeiro/contas-receber/editar", data=edit_payload, headers={"X-Requested-With": "XMLHttpRequest"})
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    updated_row = db.session.execute(text("SELECT cliente, software, total FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()
    assert updated_row.cliente == "Cliente Edit Atualizado"
    assert updated_row.software == "New Soft"
    assert updated_row.total == 950.0

    # Insert subchild and test edit_sub
    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, cnpj, contrato, software, valor, total, status, id_pai)
        VALUES ('Cliente Edit Atualizado', '22.222.222/0001-22', 'CTR-EDIT-NEW', 'Sub Soft', 200.0, 200.0, 'ABERTO', :id_pai)
    """), {"id_pai": cid})
    db.session.commit()
    sub_id = db.session.execute(text("SELECT id FROM contas_receber WHERE id_pai = :id"), {"id": cid}).fetchone()[0]

    res_sub = client.post("/financeiro/contas-receber/editar-sub", data={
        "id": str(sub_id),
        "cliente": "Cliente Edit Atualizado",
        "cnpj": "22.222.222/0001-22",
        "contrato": "CTR-EDIT-NEW",
        "unidade": "SOLLUS SP",
        "software": "Sub Soft 2",
        "data_primeira_pendencia": "2026-07-05",
        "qt_pendencias": "1",
        "valor": "300,00"
    }, headers={"X-Requested-With": "XMLHttpRequest"})
    assert res_sub.status_code == 200
    assert res_sub.get_json()["ok"] is True

    # Check recalculation of parent total (950 + 300 = 1250)
    parent_after = db.session.execute(text("SELECT total FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()
    assert parent_after.total == 1250.0


def test_contas_receber_update_info_and_historico(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, cnpj, contrato, valor, status)
        VALUES ('Cliente Info', '33.333.333/0001-33', 'CTR-INFO', 500.0, 'ABERTO')
    """))
    db.session.commit()
    cid = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Info'")).fetchone()[0]

    # Update info
    res = client.post("/financeiro/contas-receber/update-info", data={
        "id": str(cid),
        "new_info": "Contato feito com o financeiro do cliente",
    })
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    # Fetch historico
    hist_res = client.post("/financeiro/contas-receber/historico", data={"id": str(cid)})
    assert hist_res.status_code == 200
    hist_json = hist_res.get_json()
    assert hist_json["ok"] is True
    assert "Contato feito com o financeiro do cliente" in hist_json["html"]


def test_contas_receber_valor(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, valor, status)
        VALUES ('Cliente Valor', 100.0, 'ABERTO')
    """))
    db.session.commit()
    cid = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Valor'")).fetchone()[0]

    res = client.post("/financeiro/contas-receber/valor", data={
        "id": str(cid),
        "valor": "150,00",
    })
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    row = db.session.execute(text("SELECT valor, informacoes FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()
    assert row.valor == 250.0
    assert "Valor adicionado: R$ 150,00" in row.informacoes


def test_contas_receber_quitar(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, valor, status, id_pai)
        VALUES ('Cliente Quitar', 500.0, 'ABERTO', 0)
    """))
    db.session.commit()
    cid = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Quitar'")).fetchone()[0]

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, valor, status, id_pai)
        VALUES ('Cliente Quitar Sub', 200.0, 'ABERTO', :id_pai)
    """), {"id_pai": cid})
    db.session.commit()

    res = client.post("/financeiro/contas-receber/quitar", data={"id": str(cid)})
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    parent_status = db.session.execute(text("SELECT status FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()[0]
    child_status = db.session.execute(text("SELECT status FROM contas_receber WHERE id_pai = :id"), {"id": cid}).fetchone()[0]
    assert parent_status == "QUITADO"
    assert child_status == "QUITADO"


def test_contas_receber_bloqueio_cancelamento_deferimento(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, cnpj, contrato, software, data_primeira_pendencia, valor, total, status)
        VALUES ('Cliente Workflow', '44.444.444/0001-44', 'CTR-WORK', 'Sollus', '2026-06-01', 300.0, 300.0, 'ABERTO')
    """))
    db.session.commit()
    cid = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Workflow'")).fetchone()[0]

    # Solicitar Bloqueio
    with patch("modules.financeiro.blueprints.financeiro._dispatch_email_async") as mock_email:
        res_bloq = client.post("/financeiro/contas-receber/solicitar-bloqueio", data={
            "id": str(cid),
            "data_bloqueio": "2026-07-20",
        })
        assert res_bloq.status_code == 200
        assert res_bloq.get_json()["ok"] is True
        assert mock_email.called

    row_bloq = db.session.execute(text("SELECT data_bloqueio, informacoes FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()
    assert row_bloq.data_bloqueio == "2026-07-20"
    assert "solicitou o bloqueio" in row_bloq.informacoes

    # Solicitar Cancelamento
    with patch("modules.financeiro.blueprints.financeiro._dispatch_email_async") as mock_email:
        res_canc = client.post("/financeiro/contas-receber/cancelamento", data={
            "id": str(cid),
            "cancelamento": "2026-07-22",
        })
        assert res_canc.status_code == 200
        assert res_canc.get_json()["ok"] is True
        assert mock_email.called

    row_canc = db.session.execute(text("SELECT cancelamento, informacoes FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()
    assert row_canc.cancelamento == "2026-07-22"
    assert "solicitou o Cancelamento" in row_canc.informacoes

    # Deferimento
    with patch("modules.financeiro.blueprints.financeiro._dispatch_email_async") as mock_email:
        res_def = client.post("/financeiro/contas-receber/deferimento", data={
            "id": str(cid),
            "deferimento": "2026-07-25",
            "diasAtraso": "55",
        })
        assert res_def.status_code == 200
        assert res_def.get_json()["ok"] is True
        assert mock_email.called

    row_def = db.session.execute(text("SELECT status, deferimento_cancelamento, dias_atraso FROM contas_receber WHERE id = :id"), {"id": cid}).fetchone()
    assert row_def.status == "FECHADO"
    assert row_def.deferimento_cancelamento == "2026-07-25"
    assert row_def.dias_atraso == 55


def test_contas_receber_cancelados_list(client):
    login(client, username="admin_fin")

    # Insert closed and quitado accounts
    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, status) VALUES ('Cliente Fechado', 'FECHADO');
    """))
    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, status) VALUES ('Cliente Quitado', 'QUITADO');
    """))
    db.session.commit()

    res = client.get("/financeiro/contas-receber/cancelados")
    assert res.status_code == 200
    assert b"Cliente Fechado" in res.data or b"Cliente Quitado" in res.data or b"Registros" in res.data


def test_contas_receber_excluir(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, status) VALUES ('Cliente Excluir', 'ABERTO')
    """))
    db.session.commit()
    cid = db.session.execute(text("SELECT id FROM contas_receber WHERE cliente = 'Cliente Excluir'")).fetchone()[0]

    # Subchild
    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, status, id_pai) VALUES ('Cliente Excluir Sub', 'ABERTO', :id_pai)
    """), {"id_pai": cid})
    db.session.commit()

    res = client.post("/financeiro/contas-receber/excluir", data={"id": str(cid)})
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    # Verify both deleted
    count = db.session.execute(text("SELECT COUNT(id) FROM contas_receber WHERE id = :id OR id_pai = :id"), {"id": cid}).scalar()
    assert count == 0

    # Non-existent delete returns 404
    res_notfound = client.post("/financeiro/contas-receber/excluir", data={"id": "999999"})
    assert res_notfound.status_code == 404
    assert res_notfound.get_json()["ok"] is False


def test_contas_receber_export_excel(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO contas_receber (cliente, cnpj, contrato, software, valor, total, status)
        VALUES ('Cliente Export', '55.555.555/0001-55', 'CTR-EXP', 'Sollus', 1200.0, 1200.0, 'ABERTO')
    """))
    db.session.commit()

    res = client.get("/financeiro/contas-receber/export?page_mode=aberto")
    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Verify openpyxl can read the returned workbook
    wb = openpyxl.load_workbook(io.BytesIO(res.data))
    ws = wb.active
    assert ws.title == "Contas a Receber"
    assert ws.cell(row=1, column=2).value == "Cliente"
    assert ws.cell(row=2, column=2).value == "Cliente Export"


# ==============================================================================
# 4. VERIFICA EMPRESA ENDPOINT
# ==============================================================================

def test_verifica_empresa(client):
    login(client, username="admin_fin")

    db.session.execute(text("""
        INSERT INTO empresa (cliente, cnpj) VALUES ('Empresa Cadastrada SA', '66666666000166')
    """))
    db.session.commit()

    # Empty CNPJ -> empty string
    res_empty = client.get("/financeiro/verifica-empresa")
    assert res_empty.status_code == 200
    assert res_empty.data.decode("utf-8") == ""

    # Existing CNPJ -> returns name from DB
    res_db = client.get("/financeiro/verifica-empresa?cnpj=66.666.666/0001-66")
    assert res_db.status_code == 200
    assert res_db.data.decode("utf-8") == "Empresa Cadastrada SA"

    # Non-existing CNPJ (mocking requests.get fallback)
    with patch("requests.get") as mock_get:
        mock_resp = MagicMock()
        mock_resp.ok = True
        mock_resp.json.return_value = {"nome": "Empresa Externa Receita"}
        mock_get.return_value = mock_resp

        res_ext = client.get("/financeiro/verifica-empresa?cnpj=77.777.777/0001-77")
        assert res_ext.status_code == 200
        assert res_ext.data.decode("utf-8") == "Empresa Externa Receita"


# ==============================================================================
# 5. COTA DASHBOARD & COTA OPERATIONS
# ==============================================================================

def test_cota_dashboard_get(client):
    login(client, username="admin_fin")
    res = client.get("/financeiro/cota")
    assert res.status_code == 200
    assert b"cota" in res.data or b"Dashboard" in res.data or b"Faturamento" in res.data


def test_cota_adicionar_and_valor_updates(client):
    login(client, username="admin_fin")

    today_str = date.today().strftime("%Y-%m-01")

    res_add = client.post("/financeiro/cota/adicionar", data={
        "valorMes": "10.000,00",
        "dataInicio": today_str,
    }, follow_redirects=True)
    assert res_add.status_code == 200

    # Verify rows created for all units
    rows = db.session.execute(text("SELECT * FROM cota_mensal WHERE Data_Inicio = :d"), {"d": today_str}).fetchall()
    assert len(rows) == len(FINANCEIRO_UNIDADES_COTA)
    cid = rows[0].ID

    # Test update cota_valor (arrecadado)
    res_v = client.post("/financeiro/cota/valor", data={
        "id": str(cid),
        "valor": "2.500,00",
        "tipo": "adicionar",
    })
    assert res_v.status_code == 200
    assert res_v.get_json()["ok"] is True

    val_arrecadado = db.session.execute(text("SELECT Valor_Arrecadado FROM cota_mensal WHERE ID = :id"), {"id": cid}).scalar()
    assert val_arrecadado == 2500.0

    # Test update cota_valor_pago
    res_p = client.post("/financeiro/cota/valor-pago", data={
        "id": str(cid),
        "valor": "1.000,00",
        "tipo": "adicionar",
    })
    assert res_p.status_code == 200
    assert res_p.get_json()["ok"] is True

    val_pago = db.session.execute(text("SELECT Valor_Pago FROM cota_mensal WHERE ID = :id"), {"id": cid}).scalar()
    assert val_pago == 1000.0


def test_cota_fechar_mes(client):
    login(client, username="admin_fin")

    dt_str = "2026-05-01"
    db.session.execute(text("""
        INSERT INTO cota_mensal (Valor_Mes, Data_Inicio, Data_Fim, status)
        VALUES (5000.0, '2026-05-01', '2026-05-31', 'ABERTO')
    """))
    db.session.commit()

    res = client.post("/financeiro/cota/fechar-mes", data={"filtro": "05-2026"})
    assert res.status_code == 200
    assert res.get_json()["ok"] is True

    st = db.session.execute(text("SELECT status FROM cota_mensal WHERE Data_Inicio = '2026-05-01'")).scalar()
    assert st == "FECHADO"


def test_cota_trimestre_and_pdfs(client):
    login(client, username="admin_fin")

    d_start = date(2026, 1, 1)
    d_end = date(2026, 1, 31)

    db.session.execute(text("""
        INSERT INTO cota_mensal (Valor_Mes, Valor_Arrecadado, Valor_Pago, Data_Inicio, Data_Fim, Unidade_Responsavel)
        VALUES (10000.0, 8000.0, 7500.0, :s, :e, 'SOLLUS RJ')
    """), {"s": d_start, "e": d_end})
    db.session.commit()

    # Create trimestral quota
    res_tri = client.post("/financeiro/cota/trimestre", data={"dataInicio": "2026-01-01"}, follow_redirects=True)
    assert res_tri.status_code == 200

    tri_row = db.session.execute(text("SELECT * FROM cota_trimestral WHERE Data_inicio = '2026-01-01'")).fetchone()
    assert tri_row is not None
    assert tri_row.Valor_faturado == 8000.0
    assert tri_row.Valor_recebido == 7500.0

    per_row = db.session.execute(text("SELECT * FROM periodo_trimestre WHERE id_trimestre = :id"), {"id": tri_row.ID}).fetchone()
    assert per_row is not None

    # Test PDF Mensal
    with patch("modules.financeiro.blueprints.financeiro.render_proposta_html_pdf") as mock_pdf:
        mock_pdf.return_value = b"%PDF-1.4 dummy monthly content"
        res_pdf_m = client.get("/financeiro/cota/pdf-mensal?mes=1&ano=2026")
        assert res_pdf_m.status_code == 200
        assert b"%PDF" in res_pdf_m.data

    # Test PDF Trimestre
    with patch("modules.financeiro.blueprints.financeiro.render_proposta_html_pdf") as mock_pdf:
        mock_pdf.return_value = b"%PDF-1.4 dummy quarterly content"
        res_pdf_t = client.get("/financeiro/cota/pdf-trimestre?data_inicial=2026-01-01")
        assert res_pdf_t.status_code == 200
        assert b"%PDF" in res_pdf_t.data


def test_cota_enviar_email(client):
    login(client, username="admin_fin")

    with patch("modules.financeiro.blueprints.financeiro._send_email") as mock_send:
        mock_send.return_value = True
        res = client.post("/financeiro/cota/enviar-email", data={"tabela": "<table><tr><td>Dados</td></tr></table>"})
        assert res.status_code == 200
        assert res.get_json()["ok"] is True
        assert mock_send.called

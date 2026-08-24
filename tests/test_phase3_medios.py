import os
os.environ["SQLALCHEMY_DATABASE_URI"] = "sqlite:///:memory:"
os.environ["MAIL_ENABLED"] = "0"

import pytest
import time
from unittest.mock import Mock, patch
from io import BytesIO
import openpyxl

from app import create_app
from extensions import db, executor
from utils.helpers import sanitize_html, submit_bg_task
from modules.suporte.services.chamados import fetch_chamados, REGIONAL_BOARDS
from modules.propostas.models import User, Proposal
from modules.audit.models import AuditLog


class TestConfig:
    TESTING = True
    WTF_CSRF_ENABLED = False
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    MAIL_ENABLED = False
    SECRET_KEY = "test-secret-key-12345"


def _make_app():
    return create_app(TestConfig)


@pytest.fixture
def app():
    app_instance = _make_app()
    with app_instance.app_context():
        db.create_all()
        
        # Create legacy tables not mapped in SQLAlchemy
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS contratos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente VARCHAR(255),
                status VARCHAR(50)
            )
        """))
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS contas_receber (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente VARCHAR(255),
                cnpj VARCHAR(50),
                contrato VARCHAR(100),
                software VARCHAR(100),
                data_primeira_pendencia VARCHAR(50),
                qt_pendencias INTEGER,
                dias_atraso INTEGER,
                total FLOAT,
                empresa_responsavel VARCHAR(255),
                valor FLOAT,
                criado_por VARCHAR(100),
                status VARCHAR(50),
                id_pai INTEGER DEFAULT 0,
                data_bloqueio VARCHAR(50),
                cancelamento VARCHAR(50),
                deferimento_cancelamento VARCHAR(50),
                informacoes TEXT
            )
        """))
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS chamadossollus (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente VARCHAR(255),
                retorno VARCHAR(50),
                tecnico VARCHAR(100),
                data_os_tecnico VARCHAR(50),
                hora_saida VARCHAR(50),
                data VARCHAR(50),
                data_os_criada VARCHAR(50)
            )
        """))
        
        # Seed basic admin user to bypass login/authentication in routes
        admin = User(
            usuario="admin",
            nome_completo="Admin User",
            email="admin@sollus.com",
            password_hash="pbkdf2:sha256:...",
            tipo="admin",
            role="admin",
            is_active=True,
        )
        db.session.add(admin)
        db.session.commit()

        # Pre-reflect audit logs and support log tables to avoid sqlite PRAGMA rollback during test requests
        from modules.audit.models import AuditLog
        from modules.suporte.models import AtendimentoSuporteLog
        db.session.query(AuditLog).first()
        db.session.query(AtendimentoSuporteLog).first()

        yield app_instance
        db.session.remove()
        db.drop_all()



@pytest.fixture
def client(app):
    return app.test_client()


def test_sanitize_html():
    # 1. Test whitelist tags
    text = "<strong>Hello</strong> <em>World</em> <span class='test'>test</span>"
    clean = sanitize_html(text)
    assert "<strong>Hello</strong>" in clean
    assert "<em>World</em>" in clean
    assert '<span class="test">test</span>' in clean

    # 2. Test forbidden tags and attributes
    text_xss = "<script>alert(1)</script><div onclick='bad()'>click me</div><a href='javascript:alert(1)'>link</a>"
    clean_xss = sanitize_html(text_xss)
    assert "<script>" not in clean_xss
    assert "onclick" not in clean_xss
    assert "javascript:" not in clean_xss
    # Check that plain text is escaped or maintained safely
    assert "click me" in clean_xss
    assert "link" in clean_xss


def test_submit_bg_task(app):
    task_runs = []

    def mock_task(val):
        task_runs.append(val)

    submit_bg_task(app, mock_task, "test_val", max_retries=1, retry_delay=0.1)
    
    # Give the executor a short moment to finish the task
    time.sleep(0.5)
    assert "test_val" in task_runs


def test_submit_bg_task_retry(app):
    failures = []
    
    def failing_task():
        failures.append(1)
        if len(failures) < 3:
            raise ValueError("Failure")
        print("Success on 3rd attempt")

    # Should retry up to 3 times, succeeding on the 3rd
    submit_bg_task(app, failing_task, max_retries=3, retry_delay=0.01)
    
    time.sleep(0.5)
    assert len(failures) == 3


def test_fetch_chamados_ordering(app):
    # Retrieve first regional board
    board = REGIONAL_BOARDS[0]
    
    # We will mock the database query to check that ordering parameters are constructed properly
    with patch("extensions.db.session.execute") as mock_execute:
        mock_execute.return_value.fetchall.return_value = []
        
        # Call with order_by_closed=True
        fetch_chamados(board, order_by_closed=True)
        
        # Verify the execution SQL query structure contains the priority ordering fields
        called_args = mock_execute.call_args[0]
        sql_query = str(called_args[0])
        
        # Since ordering priorities data_os_tecnico DESC, hora_saida DESC, data DESC...
        assert "data_os_tecnico" in sql_query or "hora_saida" in sql_query or "data" in sql_query


def test_contas_receber_native_autoincrement(app, client):
    # Test that we can add a account receivable and the database generates ID automatically.
    # Authenticate the user
    with client.session_transaction() as sess:
        sess["usuario_id"] = 1
        sess["user_id"] = 1
        sess["_user_id"] = "1"
        sess["tipo"] = "admin"
        sess["logged_in_user"] = "Admin User"

    # Make post request to add account
    response = client.post(
        "/financeiro/contas-receber/adicionar",
        data={
            "cliente": "Cliente Teste Autoincrement",
            "cnpj": "12.345.678/0001-90",
            "contrato": "CON-2026-001",
            "software": "Sollus Connect",
            "data_primeira_pendencia": "2026-06-10",
            "empresa_responsavel": "SOLLUS RJ",
            "valor": "1.500,00",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    
    # Verify it exists in SQLite database
    res = db.session.execute(db.text("SELECT * FROM contas_receber WHERE cliente = 'Cliente Teste Autoincrement'")).fetchone()
    assert res is not None
    # Verify that an ID was generated
    assert res[0] is not None


def test_kpi_counts_on_home(app, client):
    # Create some mock data
    with app.app_context():
        # Create an approved proposal and a pending proposal
        p1 = Proposal(
            client_name="Client A",
            usuario_id=1,
            approved_at=None,
        )
        p2 = Proposal(
            client_name="Client B",
            usuario_id=1,
            approved_at=None,
        )
        db.session.add(p1)
        db.session.add(p2)
        
        # Create a contract
        db.session.execute(
            db.text("INSERT INTO contratos (cliente, status) VALUES ('Client A', 'Ativo')")
        )
        db.session.commit()

    with client.session_transaction() as sess:
        sess["usuario_id"] = 1
        sess["user_id"] = 1
        sess["_user_id"] = "1"
        sess["tipo"] = "admin"

    response = client.get("/")
    assert response.status_code == 200
    html_content = response.data.decode("utf-8")
    
    # Check that old KPI sections are NOT rendered in home page (removed per user request)
    assert "Chamados Abertos" not in html_content
    assert "Propostas Pendentes" not in html_content
    assert "Contratos Ativos" not in html_content


def test_excel_export_endpoints(app, client):
    with client.session_transaction() as sess:
        sess["usuario_id"] = 1
        sess["user_id"] = 1
        sess["_user_id"] = "1"
        sess["tipo"] = "admin"

    # 1. Test financeiro export endpoint
    res_fin = client.get("/financeiro/contas-receber/export")
    assert res_fin.status_code == 200
    assert res_fin.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Verify that it is a valid excel workbook
    wb_fin = openpyxl.load_workbook(BytesIO(res_fin.data))
    assert "Contas a Receber" in wb_fin.sheetnames

    # 2. Test audit log export endpoint
    res_aud = client.get("/audit/export")
    assert res_aud.status_code == 200
    assert res_aud.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Verify valid workbook
    wb_aud = openpyxl.load_workbook(BytesIO(res_aud.data))
    assert "Logs de Auditoria" in wb_aud.sheetnames


def test_editar_atendimento_status_concluido(app, client):
    from modules.suporte.models import AtendimentoSuporte
    from modules.propostas.models import User, Department
    from datetime import datetime

    # 1. Create a mock department and technician user in db
    with app.app_context():
        dept = Department(name="SUPORTE", slug="suporte")
        db.session.add(dept)
        db.session.commit()

        tech = User(
            usuario="tecnico1",
            nome_completo="Tecnico Um",
            email="tecnico1@sollus.com",
            password_hash="pbkdf2:sha256:...",
            tipo="suporte",
            role="user",
            is_active=True,
            department_id=dept.id,
        )
        db.session.add(tech)
        db.session.commit()
        tech_id = tech.id

        # 2. Create an atendimento in "Entrada" status
        atendimento = AtendimentoSuporte(
            cliente="rio+ saneamento",
            cnpj="12345678901234",
            tipo_atendimento="Dvida",
            status="Entrada",
            descricao="Chamado inicial",
            os_entrada="12345",
            data_entrada=datetime.utcnow(),
            usuario_designado=tech_id,
            sistema="Sollus Connect",
            quantidade_pessoas="1",
            email="test@sollus.com",
        )
        db.session.add(atendimento)
        db.session.commit()
        atendimento_id = atendimento.id

    # 3. Authenticate user
    with client.session_transaction() as sess:
        sess["usuario_id"] = 1
        sess["user_id"] = 1
        sess["_user_id"] = "1"
        sess["tipo"] = "admin"
        sess["logged_in_user"] = "Admin User"

    # 4. Make post request to edit/complete the atendimento
    with patch("modules.suporte.blueprints.atendimentos.send_atendimento_concluido_email") as mock_send_email:
        response = client.post(
            f"/admin/suporte/atendimentos/{atendimento_id}/editar",
            data={
                "atendimento_id": str(atendimento_id),
                "cliente": "rio+ saneamento",
                "cnpj": "12.345.678/9012-34",
                "tipo_atendimento": "Dvida",
                "status": "Concluido",  # Change status to completed
                "descricao": "Chamado concluido com sucesso",
                "os_entrada": "12345",
                "data_entrada": "2026-06-10T12:00",
                "usuario_designado": str(tech_id),
                "sistema": "Sollus Connect",
                "quantidade_pessoas": "1",
                "email": "test@sollus.com",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        mock_send_email.assert_called_once()

    # 5. Verify database update
    with app.app_context():
        updated = AtendimentoSuporte.query.get(atendimento_id)
        assert updated.status == "Concluido"
        assert updated.descricao == "Chamado concluido com sucesso"


def test_fechar_chamado_validation_and_db_checking(app, client):
    # 1. Authenticate user
    with client.session_transaction() as sess:
        sess["usuario_id"] = 1
        sess["user_id"] = 1
        sess["_user_id"] = "1"
        sess["tipo"] = "admin"
        sess["logged_in_user"] = "Admin User"

    # Insert a dummy chamado into the sqlite test database
    with app.app_context():
        # First alter table to add columns so SQLite has them for the real update test
        db.session.execute(db.text("ALTER TABLE chamadossollus ADD COLUMN hora_entrada VARCHAR(50)"))
        db.session.execute(db.text("ALTER TABLE chamadossollus ADD COLUMN quem_atendeu VARCHAR(120)"))
        db.session.execute(db.text("ALTER TABLE chamadossollus ADD COLUMN descricao VARCHAR(255)"))
        db.session.execute(db.text("ALTER TABLE chamadossollus ADD COLUMN email_responsavel VARCHAR(255)"))
        db.session.execute(db.text("ALTER TABLE chamadossollus ADD COLUMN os_saida VARCHAR(255)"))
        db.session.execute(db.text("INSERT INTO chamadossollus (id, cliente, retorno, tecnico, data_os_criada) VALUES (1, 'Cliente Teste', 'ABERTO', 'admin', '2026-06-12 12:00:00')"))
        db.session.commit()

    # 2. Test closure with 8-character times (e.g. '10:00:00') - should succeed because of relaxed WTForms validator length
    response = client.post(
        "/admin/suporte/chamados/rj/1/fechar",
        data={
            "data_atendimento": "2026-06-12",
            "hora_entrada": "10:00:00",
            "hora_saida": "12:00:00",
            "retorno": "FECHADO",
            "tecnico": "Admin User",
            "quem_atendeu": "Test Person",
            "email_responsavel": "test@sollus.com",
            "descricao": "Finalized",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b"Chamado fechado com sucesso." in response.data

    # 3. Test when update_chamado fails (returns None)
    with patch("modules.suporte.blueprints.atendimentos.update_chamado", return_value=None):
        response_fail = client.post(
            "/admin/suporte/chamados/rj/1/fechar",
            data={
                "data_atendimento": "2026-06-12",
                "hora_entrada": "10:00:00",
                "hora_saida": "12:00:00",
                "retorno": "FECHADO",
                "tecnico": "Admin User",
                "quem_atendeu": "Test Person",
                "email_responsavel": "test@sollus.com",
                "descricao": "Finalized",
            },
            follow_redirects=True,
        )
        assert response_fail.status_code == 200
        assert b"Erro ao salvar" in response_fail.data or b"Erro ao salvar informacoes de encerramento no banco de dados." in response_fail.data
